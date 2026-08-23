"""
Enhanced WhatsApp Reminder System for Pillay Sir's ICSE Classes
Now integrated with the new FeeInstallments system
Sends reminders based on installment status: Due and Pending only

Requirements:
- Django environment setup
- selenium
- datetime
- pandas (for logging)

Usage:
python manage.py shell < send_whatsapp_reminders_installments.py
"""

import os
import logging
import django
from datetime import date, datetime
import traceback

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fees_management_system.settings')
django.setup()

# Import Django models
from core.models import Student, FeeInstallments, CustomMessageLog
from core.whatsapp_service import send_admin_summary, send_whatsapp_to_student_contacts, cleanup_whatsapp_session

logger = logging.getLogger('whatsapp')

# Alias registration_no as registration_number on Student class
Student.registration_number = property(lambda self: self.registration_no)

# Message templates based on Pillay Sir's specific timing and messages
MESSAGE_TEMPLATES = {
    'reminder_1': """Pillay Sir's ICSE Classes
Reminder -1
Kindly, make a note that your ward's next installment is due in three days time. 
Pillay Sir.
If paid , Kindly Ignore....
""",

    'reminder_2': """Pillay Sir's ICSE Classes
Reminder -2 
Kindly, pay your ward's next installment in two days time. 
Pillay Sir.
If paid , Kindly Ignore....
""",

    'reminder_3': """Pillay Sir's ICSE Classes 
Reminder -3 
Kindly pay next installment in two days time to avoid discontinuation of your ward's class. 
Pillay Sir.
If paid , Kindly Ignore....
""",

    'last_reminder': """Pillay Sir's ICSE Classes 
Last Reminder 
Kindly pay your ward's next installment today before 8:00 p.m. without fail to avoid discontinuation. 
Pillay Sir.
If paid , Kindly Ignore....
""",

    'discontinuation': """Pillay Coaching Classes 
Regret to inform you that due to non-payment of fees, which was reminded for a long time, we are not able to pay part of your ward's fees to the subject teachers and with a heavy heart decide to discontinue teaching your ward till the fees is paid with effect from today as fees is badly needed to meet all the expenses. Feel free to contact & share for the same.
""",
    
    'birthday': """Pillay Sir's ICSE Classes
Dear {first_name},
Wishing you a very Happy Birthday!
May your day be filled with joy and success.
Best Wishes,
Pillay Sir.
"""
}

def get_message_template_and_type(installment):
    """Determine message template based on Pillay Sir's specific timing requirements"""
    today = date.today()
    days_difference = (today - installment.due_date).days  # Positive = overdue, Negative = before due
    
    # Check if installment is paid - NO REMINDERS for paid installments
    if installment.status == 'Paid':
        return None, None
    
    # Pillay Sir's specific timing logic:
    if days_difference == -3:  # 3 days before due date
        return MESSAGE_TEMPLATES['reminder_1'], 'reminder_1'
    elif days_difference == 1:  # 1 day after due date
        return MESSAGE_TEMPLATES['reminder_2'], 'reminder_2'
    elif days_difference == 4:  # 4 days after due date
        return MESSAGE_TEMPLATES['reminder_3'], 'reminder_3'
    elif days_difference == 7:  # 7 days after due date
        return MESSAGE_TEMPLATES['last_reminder'], 'last_reminder'
    elif days_difference == 10:  # 10 days after due date
        return MESSAGE_TEMPLATES['discontinuation'], 'discontinuation'
    
    # No reminder for other days
    return None, None

def format_message(template, installment=None, student=None, contact_number="Pillay Sir's ICSE Classes"):
    """Format message template - Pillay Sir's messages are pre-formatted"""
    if '{first_name}' in template and student:
        # Format birthday message with student's first name
        return template.format(first_name=student.first_name)
    # Pillay Sir's fee reminder messages are already complete and don't need formatting
    return template


def format_registration_no(reg_no):
    """Return the student's registration number as exactly 6 digits."""
    try:
        return f"{int(reg_no):06d}"
    except (TypeError, ValueError):
        return str(reg_no).strip()


def build_admin_summary(failed_registrations):
    """Build the one-per-batch admin summary from student-level failures."""
    unique_failed = sorted(set(failed_registrations))

    if not unique_failed:
        return "WhatsApp Batch Summary\n\nAll messages delivered successfully."

    summary_lines = [
        "WhatsApp Batch Summary",
        "",
        "Students where both parent numbers failed:",
        "",
    ]
    for reg_no in unique_failed:
        summary_lines.append(f"* {reg_no}")
    summary_lines.append("")
    summary_lines.append(f"Total: {len(unique_failed)}")
    return "\n".join(summary_lines)


def log_batch_result(student, result, context):
    student_id = getattr(student, 'registration_no', 'unknown')
    logger.debug(
        'WhatsApp batch debug | context=%s | student_id=%s | parent_numbers_attempted=%s | successful_numbers=%s | result_success=%s | ui_status=%s',
        context,
        student_id,
        result.attempted_numbers,
        result.successful_numbers,
        result.success,
        result.status,
    )
    print(f"Student: {student_id}")
    print(f"Parent numbers attempted: {result.attempted_numbers}")
    print(f"Successful numbers: {result.successful_numbers}")
    print(f"Delivered: {result.success}")
    print(f"UI Status: {result.status}")


def add_failed_registration(failed_registrations, reg_no, context):
    formatted_reg_no = format_registration_no(reg_no)
    failed_registrations.append(formatted_reg_no)
    logger.debug(
        'WhatsApp batch debug | context=%s | failed_registrations_added=%s | failed_registrations=%s',
        context,
        formatted_reg_no,
        failed_registrations,
    )
    print(f"Added to failed_registrations: {formatted_reg_no}")

def get_birthday_students():
    """Get students who have birthdays today"""
    today = date.today()
    birthday_students = Student.objects.filter(
        date_of_birth__month=today.month,
        date_of_birth__day=today.day
    ).select_related('parentinfo')
    print(f"Birthday students found: {len(birthday_students)}")
    return birthday_students

def send_whatsapp_message(phone_number, message):
    """Send WhatsApp message using PyWhatKit with delivery verification."""
    from core.whatsapp_service import send_whatsapp_message as send_via_service

    result = send_via_service(phone_number, message, wait_time=15)
    if not result.success:
        print(f"Error sending WhatsApp message to {phone_number}: {result.error}")
    return result.success


def send_admin_notification(message_type, student_name, phone_number, message_sent):
    """Send notification to admin about sent messages."""
    from core.whatsapp_service import send_admin_notification as notify_admin

    try:
        result = notify_admin(message_type, student_name, phone_number, message_sent)
        if result.success:
            print("  Admin notification sent successfully")
        else:
            print(f"  Failed to send admin notification: {result.error}")
        return result.success
    except Exception as e:
        print(f"  Admin notification exception (non-fatal): {e}")
        traceback.print_exc()
        return False

def get_installments_for_reminders():
    """Get installments that need reminders based on Pillay Sir's specific timing"""
    today = date.today()
    
    # Get all non-paid installments
    all_installments = FeeInstallments.objects.exclude(
        status='Paid'
    ).select_related('registration_no', 'registration_no__parentinfo')
    
    reminder_installments = []
    
    # Filter installments based on Pillay Sir's specific timing
    for installment in all_installments:
        days_difference = (today - installment.due_date).days
        
        # Check if this installment needs a reminder today
        if days_difference in [-3, 1, 4, 7, 10]:
            reminder_installments.append(installment)
            
    reminder_students = list({inst.registration_no for inst in reminder_installments})
    print(f"Reminder students found: {len(reminder_students)}")
    return reminder_installments


def log_message(installment, message_type, phone_number, status, message_text, student=None):
    """Log the message to database"""
    try:
        log_student = student or installment.registration_no
        CustomMessageLog.objects.create(
            student=log_student,
            message_text=message_text[:500],  # Truncate if too long
            sent_status=status,
            attachment=(
                f"Installment {installment.installment_no} - {message_type}"
                if installment else f"{message_type} - {phone_number}"
            )
        )
    except Exception as e:
        print(f"Error logging message: {e}")

def create_excel_log(sent_messages, failed_messages):
    """Create Excel log of sent messages"""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill
        
        # Prepare data
        all_messages = []
        
        for msg in sent_messages:
            all_messages.append({
                'Student Name': msg['student_name'],
                'Registration No': msg['reg_no'],
                'Phone Number': msg['phone_number'],
                'Installment No': msg['installment_no'],
                'Amount': msg['amount'],
                'Due Date': msg['due_date'],
                'Days Difference': msg['days_difference'],
                'Message Type': msg['message_type'],
                'Status': 'SUCCESS',
                'Timestamp': msg['timestamp']
            })
        
        for msg in failed_messages:
            all_messages.append({
                'Student Name': msg['student_name'],
                'Registration No': msg['reg_no'],
                'Phone Number': msg['phone_number'],
                'Installment No': msg['installment_no'],
                'Amount': msg['amount'],
                'Due Date': msg['due_date'],
                'Days Difference': msg['days_difference'],
                'Message Type': msg['message_type'],
                'Status': 'FAILED',
                'Timestamp': msg['timestamp']
            })
        
        if not all_messages:
            return
        
        # Create DataFrame
        df = pd.DataFrame(all_messages)
        
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'logs/whatsapp_reminders_installments_{timestamp}.xlsx'
        
        os.makedirs('logs', exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='WhatsApp Reminders', index=False)
            
            # Style the worksheet
            workbook = writer.book
            worksheet = writer.sheets['WhatsApp Reminders']
            
            # Header styling
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        
        print(f"Excel log created: {filename}")
        
    except Exception as e:
        print(f"Error creating Excel log: {e}")

def main():
    """Main function to send WhatsApp reminders and birthday messages based on Pillay Sir's timing"""
    print("Starting Pillay Sir's WhatsApp Reminder System")
    print("=" * 60)
    print("Reminder Timing:")
    print("  * 3 days before due date: Reminder -1")
    print("  * 1 day after due date: Reminder -2")
    print("  * 4 days after due date: Reminder -3")
    print("  * 7 days after due date: Last Reminder")
    print("  * 10 days after due date: Discontinuation Notice")
    print("  * Paid installments: NO reminders")
    print("Birthday Messages: Sent to students on their birthday")
    print("Admin Notifications: Copy of all messages sent to admin")
    print("=" * 60)
    
    # Update overdue statuses first
    print("Updating overdue installment statuses...")
    updated_count = FeeInstallments.update_overdue_statuses()
    print(f"Updated {updated_count} installments to 'Pending' status")
    
    # Get installments that need reminders
    installments = get_installments_for_reminders()
    
    # Get students with birthdays today
    birthday_students = get_birthday_students()
    
    if not installments and not birthday_students:
        print("No installments require reminders and no birthdays today.")
        return
    
    print(f"Found {len(installments)} installments requiring reminders")
    print(f"Found {len(birthday_students)} students with birthdays today")
    
    # Ask for confirmation
    print("\n" + "=" * 60)
    print("WHATSAPP MESSAGE DETAILS:")
    print("=" * 60)
    
    # Display birthday students first
    if birthday_students:
        print("BIRTHDAY MESSAGES:")
        print("-" * 30)
        for student in birthday_students:
            print(f"* {student.full_name} (Reg: {student.registration_no})")
            print(f"  Birthday: {student.date_of_birth.strftime('%d %B %Y')}")
            if hasattr(student, 'parentinfo'):
                parent = student.parentinfo
                if parent.father_mobile:
                    print(f"  Father: {parent.father_mobile}")
                if parent.mother_mobile:
                    print(f"  Mother: {parent.mother_mobile}")
            print()
        print()
    
    # Display fee reminders
    if installments:
        print("FEE REMINDERS:")
        print("-" * 30)
    
    for installment in installments:
        template, message_type = get_message_template_and_type(installment)
        if template and message_type:
            days_difference = (date.today() - installment.due_date).days
            timing_desc = ""
            if days_difference == -3:
                timing_desc = "3 days before due date"
            elif days_difference == 1:
                timing_desc = "1 day after due date"
            elif days_difference == 4:
                timing_desc = "4 days after due date"
            elif days_difference == 7:
                timing_desc = "7 days after due date"
            elif days_difference == 10:
                timing_desc = "10 days after due date"
            
            print(f"* {installment.registration_no.full_name} (Reg: {installment.registration_no.registration_no})")
            print(f"  Installment {installment.installment_no} - Rs.{installment.amount}")
            print(f"  Due Date: {installment.due_date.strftime('%d %B %Y')}")
            print(f"  Timing: {timing_desc} - Type: {message_type}")
            if hasattr(installment.registration_no, 'parentinfo'):
                parent = installment.registration_no.parentinfo
                if parent.father_mobile:
                    print(f"  Father: {parent.father_mobile}")
                if parent.mother_mobile:
                    print(f"  Mother: {parent.mother_mobile}")
            print()
    
    # Confirmation
    total_messages = len(installments) + len(birthday_students)
    confirm = input(f"Do you want to send these {total_messages} WhatsApp messages (reminders + birthdays)? (yes/no): ").lower().strip()
    if confirm not in ['yes', 'y']:
        print("Operation cancelled by user.")
        return
    
    # Send messages
    sent_messages = []
    failed_messages = []
    
    # Initialize failed_registrations list as requested
    failed_registrations = []
    
    print("\nSending WhatsApp messages...")
    print("=" * 60)
    
    # Send birthday messages first
    birthday_count = 0
    try:
        if birthday_students:
            print("Sending Birthday Messages...")
            print("-" * 40)
            
            for student in birthday_students:
                birthday_count += 1
                try:
                    birthday_message = format_message(MESSAGE_TEMPLATES['birthday'], student=student)
                    print(f"Processing student: {student.registration_number}")

                    result = send_whatsapp_to_student_contacts(
                        student,
                        birthday_message,
                        message_type='birthday',
                    )
                    log_batch_result(student, result, 'birthday')

                    for number in result.successful_numbers:
                        sent_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': number,
                            'installment_no': 'Birthday',
                            'amount': 'N/A',
                            'due_date': student.date_of_birth.strftime('%d %B'),
                            'days_difference': 0,
                            'message_type': 'birthday',
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        })
                        send_admin_notification('birthday', student.full_name, number, birthday_message)
                        log_message(None, 'birthday', number, 'SUCCESS', birthday_message, student=student)

                    for number in result.failed_numbers:
                        failed_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': number,
                            'installment_no': 'Birthday',
                            'amount': 'N/A',
                            'due_date': student.date_of_birth.strftime('%d %B'),
                            'days_difference': 0,
                            'message_type': 'birthday',
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'error': result.failure_reason or 'Send failed',
                        })
                        log_message(None, 'birthday', number, 'FAILED', birthday_message, student=student)

                    if not result.success and not result.failed_numbers:
                        failed_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': result.attempted_numbers[-1] if result.attempted_numbers else 'N/A',
                            'installment_no': 'Birthday',
                            'amount': 'N/A',
                            'due_date': student.date_of_birth.strftime('%d %B'),
                            'days_difference': 0,
                            'message_type': 'birthday',
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'error': result.failure_reason or 'Send failed',
                        })

                    if not result.success:
                        add_failed_registration(failed_registrations, student.registration_no, 'birthday')
                        
                except Exception as student_exc:
                    print(f"  Critical birthday error for student {student.full_name}:")
                    traceback.print_exc()
                    add_failed_registration(failed_registrations, student.registration_no, 'birthday_exception')
                    continue

            if birthday_students:
                print()
    except Exception as e:
        print("Critical error in birthday processing block:")
        traceback.print_exc()

    # Send fee reminder messages
    try:
        if installments:
            print("Sending Fee Reminders...")
            print("-" * 40)

            for i, installment in enumerate(installments, 1):
                try:
                    student = installment.registration_no
                    template, message_type = get_message_template_and_type(installment)
                    if not template or not message_type:
                        continue

                    message = format_message(template, installment)
                    if not message:
                        continue

                    print(f"Processing student: {student.registration_number}")

                    result = send_whatsapp_to_student_contacts(
                        student,
                        message,
                        message_type=message_type,
                    )
                    log_batch_result(student, result, message_type)

                    for number in result.successful_numbers:
                        sent_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': number,
                            'installment_no': installment.installment_no,
                            'amount': installment.amount,
                            'due_date': installment.due_date.strftime('%d %B %Y'),
                            'days_difference': (date.today() - installment.due_date).days,
                            'message_type': message_type,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        })
                        send_admin_notification(
                            message_type,
                            student.full_name,
                            number,
                            message,
                        )
                        log_message(installment, message_type, number, 'SUCCESS', message)

                    for number in result.failed_numbers:
                        failed_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': number,
                            'installment_no': installment.installment_no,
                            'amount': installment.amount,
                            'due_date': installment.due_date.strftime('%d %B %Y'),
                            'days_difference': (date.today() - installment.due_date).days,
                            'message_type': message_type,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'error': result.failure_reason or 'Send failed',
                        })
                        log_message(installment, message_type, number, 'FAILED', message)

                    if not result.success and not result.failed_numbers:
                        failed_messages.append({
                            'student_name': student.full_name,
                            'reg_no': student.registration_no,
                            'phone_number': result.attempted_numbers[-1] if result.attempted_numbers else 'N/A',
                            'installment_no': installment.installment_no,
                            'amount': installment.amount,
                            'due_date': installment.due_date.strftime('%d %B %Y'),
                            'days_difference': (date.today() - installment.due_date).days,
                            'message_type': message_type,
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'error': result.failure_reason or 'Send failed',
                        })

                    if not result.success:
                        add_failed_registration(failed_registrations, student.registration_no, message_type)

                except Exception as installment_exc:
                    print(f"  Critical reminder error for installment {i}:")
                    traceback.print_exc()
                    if 'installment' in locals() and installment and installment.registration_no:
                        add_failed_registration(
                            failed_registrations,
                            installment.registration_no.registration_no,
                            'reminder_exception',
                        )
                    continue
    except Exception as e:
        print("Critical error in reminder processing block:")
        traceback.print_exc()
    
    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY REPORT")
    print("=" * 60)
    print(f"Messages sent successfully: {len(sent_messages)}")
    print(f"Messages failed: {len(failed_messages)}")
    print(f"Total installments processed: {len(installments)}")
    
    # Create Excel log
    create_excel_log(sent_messages, failed_messages)
    
    # Format and send one admin summary after all birthday and reminder sends finish.
    try:
        summary_msg = build_admin_summary(failed_registrations)
        logger.debug(
            'WhatsApp batch debug | admin_summary_execution=attempted | failed_registrations=%s',
            sorted(set(failed_registrations)),
        )
        print("Admin summary execution: attempted")
        send_admin_summary(summary_msg)
    except Exception as e:
        print("Error sending admin summary:")
        traceback.print_exc()
    
    print("\nWhatsApp reminder process completed!")
    print("Check the logs folder for detailed Excel report.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("Critical error in main execution:")
        traceback.print_exc()
    finally:
        # Clean up WhatsApp session (always runs, even on error)
        print("Cleaning up WhatsApp session...")
        cleanup_whatsapp_session()
