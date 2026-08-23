"""
WhatsApp Reminder System for Pillay Sir's ICSE Classes
Automatically sends WhatsApp reminders to parents for pending fees

Requirements:
- pandas
- pymysql
- pywhatkit
- datetime
- time
- openpyxl

Usage on Windows:
python send_whatsapp_reminders.py
"""

# Fix encoding issues on Windows
import sys
import os

# Set UTF-8 encoding for Windows
if sys.platform.startswith('win'):
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import pandas as pd
import pymysql
import pywhatkit as pwk
import datetime
import time
import os
import csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Root',
    'database': 'pclasses',
    'charset': 'utf8mb4'
}

# Your WhatsApp number for confirmations
MY_NUMBER = '+919822574252'  # Replace with your actual number

# Message templates based on days before/after due date
MESSAGE_TEMPLATES = {
    'reminder_1': """Pillay Sir's ICSE Classes
Reminder -1
Kindly, make a note that your ward's next installment is due in three days time.
Pillay Sir.""",
    
    'reminder_2': """Pillay Sir's ICSE Classes
Reminder -2
Kindly, pay your ward's next installment in two days time.
Pillay Sir.""",
    
    'reminder_3': """Pillay Sir's ICSE Classes
Reminder -3
Kindly pay next installment in two days time to avoid discontinuation of your ward's class.
Pillay Sir.""",
    
    'last_reminder': """Pillay Sir's ICSE Classes
Last Reminder
Kindly pay your ward's next installment today before 8:00 p.m. without fail to avoid discontinuation.
Pillay Sir.""",
    
    'discontinuation': """Pillay Coaching Classes
Regret to inform you that due to non-payment of fees, which was reminded for a long time, we are not able to pay part of your ward's fees to the subject teachers and with a heavy heart decide to discontinue teaching your ward till the fees is paid with effect from today as fees is badly needed to meet all the expenses.
Feel free to contact & share for the same. 🙏🏻""",
    
    'birthday_wish': """🎉 Pillay Sir's ICSE Classes 🎉
Dear {First_Name},
Wishing you a very Happy Birthday! 🎂
May your day be filled with joy and success.
Best Wishes,
Pillay Sir."""
}

# Create logs directory if it doesn't exist
LOGS_DIR = Path(__file__).parent / 'logs'
LOGS_DIR.mkdir(exist_ok=True)
LOG_FILE = LOGS_DIR / 'whatsapp_log.csv'
EXCEL_LOG_FILE = LOGS_DIR / f'whatsapp_detailed_log_{datetime.date.today().strftime("%Y%m%d")}.xlsx'

def setup_logging():
    """Initialize the CSV log file with headers if it doesn't exist"""
    if not LOG_FILE.exists():
        with open(LOG_FILE, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(['Timestamp', 'Student_Name', 'Registration_No', 'Mobile', 'Message_Status', 'Message_Type', 'Days_Difference', 'Due_Date', 'Fees_Amount', 'Error_Message'])

def log_message(student_name, registration_no, mobile, status, fees_amount, message_type='', days_diff=0, due_date='', error_msg=''):
    """Log WhatsApp message attempt to CSV file"""
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with open(LOG_FILE, 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow([timestamp, student_name, registration_no, mobile, status, message_type, days_diff, due_date, fees_amount, error_msg])


def create_excel_log(birthday_data, reminder_data, custom_data=None):
    """Create detailed Excel log with all WhatsApp message data"""
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for different message types
        if not birthday_data.empty:
            create_birthday_sheet(wb, birthday_data)
        
        if not reminder_data.empty:
            create_reminder_sheet(wb, reminder_data)
        
        # Create summary sheet
        create_summary_sheet(wb, birthday_data, reminder_data, custom_data)
        
        # Save the workbook
        wb.save(EXCEL_LOG_FILE)
        print(f"📊 Excel log saved: {EXCEL_LOG_FILE}")
        
    except Exception as e:
        print(f"❌ Error creating Excel log: {e}")


def create_birthday_sheet(wb, birthday_data):
    """Create birthday wishes sheet"""
    ws = wb.create_sheet("Birthday Wishes")
    
    # Headers
    headers = ['Student Name', 'Registration No', 'Date of Birth', 'Father Mobile', 'Status', 'Timestamp']
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for _, student in birthday_data.iterrows():
        status = 'Sent' if pd.notna(student.get('Father_Mobile')) or pd.notna(student.get('Mother_Mobile')) else 'Failed - No Mobile'
        ws.append([
            f"{student['First_Name']} {student['Last_Name']}",
            student['Registration_No'],
            student['Date_of_Birth'],
            student.get('Father_Mobile', 'N/A'),
            status,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])


def create_reminder_sheet(wb, reminder_data):
    """Create fee reminders sheet"""
    ws = wb.create_sheet("Fee Reminders")
    
    # Headers
    headers = ['Student Name', 'Registration No', 'Course', 'Mobile', 'Message Type', 'Days Overdue', 'Fees Remaining', 'Status', 'Timestamp']
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for _, student in reminder_data.iterrows():
        status = 'Sent' if pd.notna(student.get('Father_Mobile')) else 'Failed - No Mobile'
        ws.append([
            f"{student['First_Name']} {student['Last_Name']}",
            student['Registration_No'],
            student.get('Enrolled_Course', 'N/A'),
            student.get('Father_Mobile', 'N/A'),
            student.get('message_type', 'N/A'),
            student.get('days_difference', 0),
            f"₹{student.get('Fees_Remaining', 0):,.0f}",
            status,
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])


def create_summary_sheet(wb, birthday_data, reminder_data, custom_data):
    """Create summary sheet with statistics"""
    ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    
    # Title
    ws['A1'] = "WhatsApp Messages Summary Report"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    # Date and time
    ws['A3'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(bold=True)
    
    # Statistics
    birthday_count = len(birthday_data) if not birthday_data.empty else 0
    reminder_count = len(reminder_data) if not reminder_data.empty else 0
    custom_count = len(custom_data) if custom_data is not None and not custom_data.empty else 0
    
    ws['A5'] = "MESSAGE STATISTICS"
    ws['A5'].font = Font(bold=True, size=14)
    
    ws['A7'] = "🎂 Birthday Wishes:"
    ws['B7'] = birthday_count
    
    ws['A8'] = "💰 Fee Reminders:"
    ws['B8'] = reminder_count
    
    ws['A9'] = "📱 Custom Messages:"
    ws['B9'] = custom_count
    
    ws['A11'] = "Total Messages:"
    ws['B11'] = birthday_count + reminder_count + custom_count
    ws['B11'].font = Font(bold=True)


def get_students_data():
    """Fetch student data from MySQL database using pandas"""
    try:
        # Create database connection
        connection = pymysql.connect(**DB_CONFIG)
        
        # SQL query to get students with pending fees and due dates
        query = """
        SELECT 
            s.Registration_No,
            s.First_Name,
            s.Last_Name,
            s.Date_of_Birth,
            p.Father_Name,
            p.Father_Mobile,
            p.Mother_Mobile,
            f.Total_Fees,
            f.Fees_Remaining,
            f.Fees_Per_Installment,
            f.first_installment_date,
            f.second_installment_date,
            f.third_installment_date,
            a.Enrolled_Course
        FROM Student s
        LEFT JOIN Parent_Info p ON s.Registration_No = p.Registration_No
        LEFT JOIN Fee_Details f ON s.Registration_No = f.Registration_No
        LEFT JOIN Academic_Info a ON s.Registration_No = a.Registration_No
        WHERE f.Fees_Remaining > 0
        """
        
        # Load data into pandas DataFrame
        df = pd.read_sql(query, connection)
        connection.close()
        
        print(f"[SUCCESS] Loaded {len(df)} students with pending fees from database")
        return df
        
    except Exception as e:
        print(f"[ERROR] Error connecting to database: {e}")
        return pd.DataFrame()

def get_next_due_date(student):
    """Get the next upcoming due date for a student"""
    today = datetime.date.today()
    installment_dates = [
        student['first_installment_date'],
        student['second_installment_date'], 
        student['third_installment_date']
    ]
    
    # Filter out null dates and sort
    valid_dates = [date for date in installment_dates if pd.notna(date)]
    
    if not valid_dates:
        return None
    
    # Find the next due date (closest to today, including past dates)
    valid_dates.sort()
    
    # First, check for any future dates
    future_dates = [date for date in valid_dates if date >= today]
    if future_dates:
        return future_dates[0]
    
    # If no future dates, return the most recent past date
    return valid_dates[-1]

def calculate_days_difference(due_date):
    """Calculate days difference from today to due date (negative for overdue)"""
    if pd.isna(due_date) or due_date is None:
        return None
    
    today = datetime.date.today()
    return (due_date - today).days

def get_message_type_and_template(days_diff):
    """Determine message type based on days difference"""
    if days_diff is None:
        return None, None  # No message if no due date
    
    # New 5-level reminder schedule
    if days_diff == -3:  # 3 days before due date
        return 'reminder_1', MESSAGE_TEMPLATES['reminder_1']
    elif days_diff == 1:  # 1 day after due date
        return 'reminder_2', MESSAGE_TEMPLATES['reminder_2']
    elif days_diff == 4:  # 4 days after due date
        return 'reminder_3', MESSAGE_TEMPLATES['reminder_3']
    elif days_diff == 7:  # 7 days after due date
        return 'last_reminder', MESSAGE_TEMPLATES['last_reminder']
    elif days_diff == 10:  # 10 days after due date
        return 'discontinuation', MESSAGE_TEMPLATES['discontinuation']
    else:
        # For other cases, don't send message
        return None, None

def filter_students_for_reminders(df):
    """Filter students who need reminders based on their due dates"""
    if df.empty:
        return df
    
    # Create a list to store students who need reminders
    reminder_students = []
    
    for _, student in df.iterrows():
        due_date = get_next_due_date(student)
        days_diff = calculate_days_difference(due_date)
        
        # Check if student needs a reminder (2 days before to 2+ days after)
        message_type, message_template = get_message_type_and_template(days_diff)
        
        if message_type is not None:
            # Add additional fields to student data
            student_dict = student.to_dict()
            student_dict['next_due_date'] = due_date
            student_dict['days_difference'] = days_diff
            student_dict['message_type'] = message_type
            student_dict['message_template'] = message_template
            reminder_students.append(student_dict)
    
    if not reminder_students:
        print("[INFO] No students need reminders at this time")
        return pd.DataFrame()
    
    reminder_df = pd.DataFrame(reminder_students)
    print(f"[INFO] Found {len(reminder_df)} students who need reminders")
    return reminder_df

def format_mobile_number(mobile):
    """Format mobile number to include country code"""
    if pd.isna(mobile) or mobile == '' or mobile is None:
        return None
    
    mobile = str(mobile).strip()
    
    # Remove any non-digit characters except +
    mobile = ''.join(c for c in mobile if c.isdigit() or c == '+')
    
    # Add country code if not present
    if not mobile.startswith('+91'):
        if mobile.startswith('91'):
            mobile = '+' + mobile
        elif mobile.startswith('0'):
            mobile = '+91' + mobile[1:]
        else:
            mobile = '+91' + mobile
    
    # Validate length (should be +91 followed by 10 digits)
    if len(mobile) == 13 and mobile.startswith('+91'):
        return mobile
    
    return None

def send_whatsapp_message(mobile, message, student_name, registration_no, fees_amount, message_type='', days_diff=0, due_date=''):
    """Send WhatsApp message using pywhatkit"""
    try:
        # Format mobile number
        formatted_mobile = format_mobile_number(mobile)
        
        if not formatted_mobile:
            log_message(student_name, registration_no, mobile, 'FAILED', fees_amount, error_msg='Invalid mobile number format')
            print(f"❌ Invalid mobile number for {student_name}: {mobile}")
            return False
        
        print(f"📱 Sending WhatsApp to {student_name} ({formatted_mobile})...")
        
        # Send message instantly (requires WhatsApp Web to be open)
        pwk.sendwhatmsg_instantly(formatted_mobile, message, wait_time=15, tab_close=True)
        
        # Log success
        log_message(student_name, registration_no, formatted_mobile, 'SUCCESS', fees_amount, message_type, days_diff, due_date)
        print(f"✅ Message sent to {student_name} - {message_type} ({days_diff} days)")
        
        # Send confirmation to your number
        confirmation_msg = f"""✅ Reminder Sent
Student: {student_name}
Registration No: {registration_no}
Mobile: {formatted_mobile}
Message Type: {message_type}
Days to Due: {days_diff}
Due Date: {due_date}
Amount: ₹{fees_amount}
Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"""
        
        time.sleep(5)  # Wait before sending confirmation
        pwk.sendwhatmsg_instantly(MY_NUMBER, confirmation_msg, wait_time=10, tab_close=True)
        
        return True
        
    except Exception as e:
        error_msg = str(e)
        log_message(student_name, registration_no, mobile, 'FAILED', fees_amount, message_type, days_diff, due_date, error_msg)
        print(f"❌ Failed to send message to {student_name}: {error_msg}")
        return False


def get_birthday_students():
    """Get students whose birthday is today"""
    try:
        # Create database connection
        connection = pymysql.connect(**DB_CONFIG)
        
        # SQL query to get students with birthday today
        today = datetime.date.today()
        query = """
        SELECT 
            s.Registration_No,
            s.First_Name,
            s.Last_Name,
            s.Date_of_Birth,
            p.Father_Name,
            p.Father_Mobile,
            p.Mother_Mobile
        FROM Student s
        LEFT JOIN Parent_Info p ON s.Registration_No = p.Registration_No
        WHERE MONTH(s.Date_of_Birth) = %s AND DAY(s.Date_of_Birth) = %s
        """
        
        # Load data into pandas DataFrame
        df = pd.read_sql(query, connection, params=[today.month, today.day])
        connection.close()
        
        print(f"🎂 Found {len(df)} students with birthday today")
        return df
        
    except Exception as e:
        print(f"❌ Error getting birthday students: {e}")
        return pd.DataFrame()


def send_birthday_wishes():
    """Send birthday wishes to students whose birthday is today"""
    print("🎂 Checking for birthday students...")
    
    birthday_students = get_birthday_students()
    
    if birthday_students.empty:
        print("✅ No birthdays today")
        return 0, 0
    
    sent_count = 0
    failed_count = 0
    
    for _, student in birthday_students.iterrows():
        try:
            # Personalize birthday message
            message = MESSAGE_TEMPLATES['birthday_wish'].format(
                First_Name=student['First_Name']
            )
            
            # Try father's mobile first, then mother's
            mobile = student['Father_Mobile'] if pd.notna(student['Father_Mobile']) else student['Mother_Mobile']
            
            if pd.isna(mobile):
                print(f"⚠️ No mobile number found for {student['First_Name']} {student['Last_Name']}")
                log_message(
                    f"{student['First_Name']} {student['Last_Name']}", 
                    student['Registration_No'], 
                    'N/A', 
                    'FAILED', 
                    0,
                    'birthday_wish',
                    0,
                    datetime.date.today(),
                    'No mobile number available'
                )
                failed_count += 1
                continue
            
            # Send birthday message
            if send_whatsapp_message(
                mobile, 
                message, 
                f"{student['First_Name']} {student['Last_Name']}", 
                student['Registration_No'],
                0,  # No fees amount for birthday
                'birthday_wish',
                0,  # No days difference for birthday
                datetime.date.today()
            ):
                sent_count += 1
            else:
                failed_count += 1
            
            # Wait between messages
            print("⏳ Waiting 15 seconds before next birthday message...")
            time.sleep(15)
            
        except Exception as e:
            print(f"❌ Failed to send birthday message to {student['First_Name']} {student['Last_Name']}: {e}")
            failed_count += 1
    
    return sent_count, failed_count


def prepare_detailed_summary(birthday_sent, birthday_failed, reminder_sent, reminder_failed, reminder_students):
    """Prepare detailed summary with student names and message types"""
    summary_parts = []
    
    # Birthday wishes summary
    if birthday_sent > 0 or birthday_failed > 0:
        birthday_students = get_birthday_students()
        if not birthday_students.empty:
            summary_parts.append("🎂 BIRTHDAY MESSAGES:")
            for _, student in birthday_students.iterrows():
                name = f"{student['First_Name']} {student['Last_Name']}"
                summary_parts.append(f"  🎉 {name}")
    
    # Fee reminders summary by type
    if reminder_sent > 0:
        summary_parts.append("\n💰 FEE REMINDER BREAKDOWN:")
        
        # Group students by message type
        message_groups = {
            'reminder_1': [],
            'reminder_2': [], 
            'reminder_3': [],
            'last_reminder': [],
            'discontinuation': []
        }
        
        for _, student in reminder_students.iterrows():
            name = f"{student['First_Name']} {student['Last_Name']}"
            msg_type = student.get('message_type', 'unknown')
            days = student.get('days_difference', 0)
            amount = student.get('Fees_Remaining', 0)
            
            if msg_type in message_groups:
                message_groups[msg_type].append(f"  • {name} (₹{amount:,.0f}, {days}d)")
        
        # Add each group to summary
        type_labels = {
            'reminder_1': '📢 REMINDER 1 (-3 days):',
            'reminder_2': '⚠️ REMINDER 2 (+1 day):',
            'reminder_3': '🚨 REMINDER 3 (+4 days):',
            'last_reminder': '🔴 LAST REMINDER (+7 days):',
            'discontinuation': '❌ DISCONTINUATION (+10 days):'
        }
        
        for msg_type, students in message_groups.items():
            if students:
                summary_parts.append(f"\n{type_labels.get(msg_type, msg_type.upper())}:")
                summary_parts.extend(students)
    
    return '\n'.join(summary_parts) if summary_parts else "No detailed messages to report."


def send_reminders():
    """Main function to send WhatsApp reminders"""
    print("[START] Starting WhatsApp Reminder System...")
    print(f"[DATE] Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Setup logging
    setup_logging()
    
    # First, send birthday wishes
    print("\n[BIRTHDAY] === BIRTHDAY WISHES ===")
    birthday_sent, birthday_failed = send_birthday_wishes()
    
    # Then, send fee reminders
    print("\n[REMINDERS] === FEE REMINDERS ===")
    
    # Get student data (only students with pending fees)
    students_df = get_students_data()
    
    if students_df.empty:
        print("[ERROR] No student data found. Exiting.")
        return
    
    # Filter students who need reminders
    reminder_students = filter_students_for_reminders(students_df)
    
    if reminder_students.empty:
        print("[INFO] No students need reminders at this time.")
        return
    
    # Send reminders
    sent_count = 0
    failed_count = 0
    
    for _, student in reminder_students.iterrows():
        # Get the pre-determined message template
        message = student['message_template']
        
        # Try father's mobile first, then mother's
        mobile = student['Father_Mobile'] if pd.notna(student['Father_Mobile']) else student['Mother_Mobile']
        
        if pd.isna(mobile):
            print(f"⚠️ No mobile number found for {student['First_Name']} {student['Last_Name']}")
            log_message(
                f"{student['First_Name']} {student['Last_Name']}", 
                student['Registration_No'], 
                'N/A', 
                'FAILED', 
                student['Fees_Remaining'],
                student['message_type'],
                student['days_difference'],
                student['next_due_date'],
                'No mobile number available'
            )
            failed_count += 1
            continue
        
        # Send message
        if send_whatsapp_message(
            mobile, 
            message, 
            f"{student['First_Name']} {student['Last_Name']}", 
            student['Registration_No'],
            student['Fees_Remaining'],
            student['message_type'],
            student['days_difference'],
            student['next_due_date']
        ):
            sent_count += 1
        else:
            failed_count += 1
        
        # Wait between messages to avoid rate limiting
        print("⏳ Waiting 20 seconds before next message...")
        time.sleep(20)
    
    # Final summary
    print(f"\n📊 Summary:")
    print(f"🎂 Birthday wishes sent: {birthday_sent}")
    print(f"🎂 Birthday wishes failed: {birthday_failed}")
    print(f"💰 Fee reminders sent: {sent_count}")
    print(f"💰 Fee reminders failed: {failed_count}")
    print(f"📝 Log file: {LOG_FILE}")
    
    # Prepare detailed summary with student names and message types
    detailed_summary = prepare_detailed_summary(birthday_sent, birthday_failed, sent_count, failed_count, reminder_students)
    
    # Create Excel log with all data
    try:
        birthday_students_df = get_birthday_students()
        create_excel_log(birthday_students_df, reminder_students)
        print(f"📊 Excel log created: {EXCEL_LOG_FILE}")
    except Exception as e:
        print(f"⚠️ Could not create Excel log: {e}")
    
    # Send final summary to your number
    try:
        summary_msg = f"""📊 WhatsApp Summary - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

🎂 BIRTHDAY WISHES:
✅ Sent: {birthday_sent}
❌ Failed: {birthday_failed}

💰 FEE REMINDERS:
✅ Sent: {sent_count}
❌ Failed: {failed_count}
📝 Students processed: {len(reminder_students)}

{detailed_summary}

Pillay Sir's ICSE Classes - Automated System"""
        
        time.sleep(10)
        pwk.sendwhatmsg_instantly(MY_NUMBER, summary_msg, wait_time=10, tab_close=True)
    except Exception as e:
        print(f"⚠️ Could not send summary message: {e}")

if __name__ == "__main__":
    send_reminders()
