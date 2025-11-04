from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q, Count, F
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.template.loader import get_template
from twilio.rest import Client
from django.conf import settings
import json
import subprocess
import os
from datetime import datetime, timedelta
from .models import Student, ParentInfo, Branch, AcademicInfo, FeeDetails, CustomMessageLog, FeeInstallments, WhatsAppMessageAnalytics, MessageAnalyticsSummary
from .forms import StudentRegistrationForm, ParentInfoForm, AcademicInfoForm, FeeDetailsForm, FeePaymentForm

# Additional imports for reports and analytics
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import io
import base64
import pandas as pd
from django.db import models


def dashboard(request):
    """Dashboard view with statistics and recent activities"""
    # Calculate statistics
    total_students = Student.objects.count()
    
    # Calculate fees more accurately
    fee_aggregates = FeeDetails.objects.aggregate(
        total_fees=Sum('total_fees'),
        total_remaining=Sum('fees_remaining')
    )
    
    total_fees = fee_aggregates['total_fees'] or 0
    pending_fees = fee_aggregates['total_remaining'] or 0
    total_fees_collected = total_fees - pending_fees
    
    # Count overdue students (students with remaining fees > 0)
    overdue_students = FeeDetails.objects.filter(fees_remaining__gt=0).count()
    
    # Ensure non-negative values
    if total_fees_collected < 0:
        total_fees_collected = 0
    if pending_fees < 0:
        pending_fees = 0
    
    # Calculate installment statistics
    try:
        installment_stats = {
            'total_installments': FeeInstallments.objects.count(),
            'due_installments': FeeInstallments.objects.filter(status='Due').count(),
            'pending_installments': FeeInstallments.objects.filter(status='Pending').count(),
            'paid_installments': FeeInstallments.objects.filter(status='Paid').count(),
        }
        
        # Calculate reminders needed today
        from datetime import date
        today = date.today()
        reminders_today = 0
        
        all_installments = FeeInstallments.objects.exclude(status='Paid')
        for installment in all_installments:
            days_difference = (today - installment.due_date).days
            if days_difference in [-3, 1, 4, 7, 10]:
                reminders_today += 1
        
        installment_stats['reminders_today'] = reminders_today
        
    except Exception as e:
        # If FeeInstallments table doesn't exist or has issues
        installment_stats = {
            'total_installments': 0,
            'due_installments': 0,
            'pending_installments': 0,
            'paid_installments': 0,
            'reminders_today': 0,
        }
    
    # Get recent students (last 10) - simplified for now
    recent_students = Student.objects.order_by('-date_of_admission')[:10]
    
    context = {
        'total_students': total_students,
        'total_fees_collected': total_fees_collected,
        'pending_fees': pending_fees,
        'overdue_students': overdue_students,
        'recent_students': recent_students,
        'installment_stats': installment_stats,
    }
    
    # Debug: Print context values
    print(f"Context values: students={total_students}, collected={total_fees_collected}, pending={pending_fees}")
    
    return render(request, 'core/dashboard.html', context)


@login_required
def student_registration(request):
    """Student registration view"""
    print(f"=== REGISTRATION VIEW CALLED ===")
    print(f"Method: {request.method}")
    print(f"User: {request.user}")
    
    if request.method == 'POST':
        try:
            # Debug: Print all POST data
            print("=== REGISTRATION DEBUG ===")
            print("POST Data:", dict(request.POST))
            print("Headers:", dict(request.headers))
            
            # Validate required fields
            required_fields = [
                'registration_no', 'date_of_admission', 'first_name', 'last_name',
                'gender', 'date_of_birth', 'father_name', 'mother_name',
                'enrolled_course', 'branch_code', 'total_fees', 'number_of_installments',
                'fees_per_installment', 'fees_remaining'
            ]
            
            missing_fields = []
            for field in required_fields:
                if not request.POST.get(field):
                    missing_fields.append(field)
            
            if missing_fields:
                error_msg = f'Missing required fields: {", ".join(missing_fields)}'
                print("ERROR:", error_msg)
                return JsonResponse({
                    'success': False,
                    'message': error_msg
                })
            
            # Check if registration number already exists
            if Student.objects.filter(registration_no=request.POST['registration_no']).exists():
                error_msg = f'Student with registration number {request.POST["registration_no"]} already exists'
                print("ERROR:", error_msg)
                return JsonResponse({
                    'success': False,
                    'message': error_msg
                })
            
            # Validate branch exists
            try:
                branch = Branch.objects.get(branch_code=request.POST['branch_code'])
                print("Branch found:", branch)
            except Branch.DoesNotExist:
                error_msg = f'Branch with code {request.POST["branch_code"]} does not exist'
                print("ERROR:", error_msg)
                return JsonResponse({
                    'success': False,
                    'message': error_msg
                })
            
            print("Creating student...")
            # Create student
            student = Student.objects.create(
                registration_no=int(request.POST['registration_no']),
                date_of_admission=request.POST['date_of_admission'],
                first_name=request.POST['first_name'],
                middle_name=request.POST.get('middle_name', ''),
                last_name=request.POST['last_name'],
                gender=request.POST['gender'],
                date_of_birth=request.POST['date_of_birth'],
                address=request.POST.get('address', '')
            )
            print("Student created:", student)
            
            print("Creating parent info...")
            # Create parent info
            parent_info = ParentInfo.objects.create(
                registration_no=student,
                father_name=request.POST['father_name'],
                mother_name=request.POST['mother_name'],
                father_occupation=request.POST.get('father_occupation', ''),
                mother_occupation=request.POST.get('mother_occupation', ''),
                father_mobile=request.POST.get('father_mobile', ''),
                mother_mobile=request.POST.get('mother_mobile', ''),
                alternate_mobile=request.POST.get('alternate_mobile', '')
            )
            print("Parent info created:", parent_info)
            
            print("Creating academic info...")
            # Create academic info
            academic_info = AcademicInfo.objects.create(
                registration_no=student,
                enrolled_course=request.POST['enrolled_course'],
                branch_code=branch,
                percentage_previous_exam=float(request.POST['percentage_previous_exam']) if request.POST.get('percentage_previous_exam') else None,
                school_name=request.POST.get('school_name', '')
            )
            print("Academic info created:", academic_info)
            
            print("Creating fee details...")
            # Create fee details with installment information
            total_fees = float(request.POST['total_fees'])
            fees_remaining = float(request.POST.get('fees_remaining', 0))
            
            # Safety check: If fees_remaining is 0 or not provided, set it to total_fees
            # (assuming no payments have been made during registration)
            if fees_remaining <= 0:
                fees_remaining = total_fees
            
            fee_details_data = {
                'registration_no': student,
                'total_fees': total_fees,
                'number_of_installments': int(request.POST['number_of_installments']),
                'fees_per_installment': float(request.POST['fees_per_installment']),
                'fees_remaining': fees_remaining
            }
            
            # Add installment details if provided (first 3 static fields)
            if request.POST.get('first_installment'):
                fee_details_data['first_installment'] = float(request.POST['first_installment'])
            if request.POST.get('first_installment_date'):
                fee_details_data['first_installment_date'] = request.POST['first_installment_date']
                
            if request.POST.get('second_installment'):
                fee_details_data['second_installment'] = float(request.POST['second_installment'])
            if request.POST.get('second_installment_date'):
                fee_details_data['second_installment_date'] = request.POST['second_installment_date']
                
            if request.POST.get('third_installment'):
                fee_details_data['third_installment'] = float(request.POST['third_installment'])
            if request.POST.get('third_installment_date'):
                fee_details_data['third_installment_date'] = request.POST['third_installment_date']
            
            fee_details = FeeDetails.objects.create(**fee_details_data)
            print("Fee details created:", fee_details)
            
            # Create installments in the new FeeInstallments table if multiple installments
            installments_created = []
            try:
                from core.models import FeeInstallments
                from datetime import datetime
                
                num_installments = int(request.POST['number_of_installments'])
                if num_installments > 1:
                    print(f"Creating {num_installments} installments...")
                    
                    # Create installments based on provided data
                    for i in range(1, num_installments + 1):
                        amount = None
                        due_date = None
                        
                        # Handle first 3 installments (static fields)
                        if i == 1:
                            amount = request.POST.get('first_installment')
                            due_date = request.POST.get('first_installment_date')
                        elif i == 2:
                            amount = request.POST.get('second_installment')
                            due_date = request.POST.get('second_installment_date')
                        elif i == 3:
                            amount = request.POST.get('third_installment')
                            due_date = request.POST.get('third_installment_date')
                        else:
                            # Handle dynamic installments (4+)
                            amount = request.POST.get(f'installment_{i}_amount')
                            due_date = request.POST.get(f'installment_{i}_date')
                        
                        if amount and due_date:
                            installment = FeeInstallments.objects.create(
                                registration_no=student,
                                installment_no=i,
                                amount=float(amount),
                                due_date=due_date,
                                status='Due'
                            )
                            installments_created.append({
                                'installment_no': i,
                                'amount': float(amount),
                                'due_date': due_date,
                                'status': 'Due'
                            })
                            print(f"Installment {i} created: ₹{amount} due on {due_date}")
                        else:
                            print(f"Warning: Installment {i} data missing - amount: {amount}, date: {due_date}")
                        
            except Exception as installment_error:
                print(f"Warning: Could not create installments: {installment_error}")
            
            # Prepare response with database details
            database_details = {
                'student_id': student.registration_no,
                'student_name': student.full_name,
                'parent_info_created': bool(parent_info),
                'academic_info_created': bool(academic_info),
                'fee_details_created': bool(fee_details),
                'installments_created': len(installments_created),
                'installment_details': installments_created
            }
            
            print("Registration completed successfully!")
            print("Database Details:", database_details)
            
            # Check if debug mode is requested
            debug_mode = request.POST.get('debug_mode') == 'true'
            
            response_data = {
                'success': True,
                'message': 'Student registered successfully!',
                'registration_no': student.registration_no,
                'database_status': 'Successfully saved to MySQL database',
                'redirect_url': '/dashboard/',
                'student_count': Student.objects.count()
            }
            
            if debug_mode:
                response_data['database_details'] = database_details
                response_data['tables_updated'] = [
                    'students (Student)',
                    'parent_info (ParentInfo)', 
                    'academic_info (AcademicInfo)',
                    'fee_details (FeeDetails)',
                    f'fee_installments ({len(installments_created)} records)'
                ]
            
            return JsonResponse(response_data)
                
        except Exception as e:
            import traceback
            error_msg = f'Registration failed: {str(e)}'
            print("EXCEPTION:", error_msg)
            print("TRACEBACK:", traceback.format_exc())
            
            return JsonResponse({
                'success': False,
                'message': error_msg
            })
    
    # Get all branches for the form
    branches = Branch.objects.all()
    print("Available branches:", list(branches.values()))
    context = {'branches': branches}
    return render(request, 'core/registration.html', context)


@login_required
def students_list(request):
    """List all students with advanced filtering"""
    students = Student.objects.select_related('academicinfo__branch_code', 'parentinfo', 'feedetails').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(registration_no__icontains=search_query) |
            Q(parentinfo__father_name__icontains=search_query) |
            Q(parentinfo__mother_name__icontains=search_query) |
            Q(academicinfo__enrolled_course__icontains=search_query)
        )
    
    # Filter by course
    course_filter = request.GET.get('course', '')
    if course_filter:
        students = students.filter(academicinfo__enrolled_course__icontains=course_filter)
    
    # Filter by branch
    branch_filter = request.GET.get('branch', '')
    if branch_filter:
        students = students.filter(academicinfo__branch_code__branch_code=branch_filter)
    
    # Filter by payment status
    payment_status = request.GET.get('payment_status', '')
    if payment_status == 'paid':
        students = students.filter(feedetails__fees_remaining=0)
    elif payment_status == 'pending':
        students = students.filter(feedetails__fees_remaining__gt=0)
    elif payment_status == 'overdue':
        students = students.filter(feedetails__fees_remaining__gt=0)  # Can add date logic later
    
    # Filter by gender
    gender_filter = request.GET.get('gender', '')
    if gender_filter:
        students = students.filter(gender=gender_filter)
    
    # Sorting
    sort_by = request.GET.get('sort', 'registration_no')
    sort_order = request.GET.get('order', 'asc')
    
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    
    valid_sort_fields = ['registration_no', 'first_name', 'last_name', 'date_of_admission', 'feedetails__fees_remaining']
    if sort_by.lstrip('-') in valid_sort_fields:
        students = students.order_by(sort_by)
    
    # Get filter options for dropdowns
    courses = AcademicInfo.objects.values_list('enrolled_course', flat=True).distinct()
    branches = Branch.objects.all()
    
    context = {
        'students': students,
        'search_query': search_query,
        'course_filter': course_filter,
        'branch_filter': branch_filter,
        'payment_status': payment_status,
        'gender_filter': gender_filter,
        'sort_by': sort_by.lstrip('-'),
        'sort_order': sort_order,
        'courses': courses,
        'branches': branches,
        'total_students': students.count(),
    }
    return render(request, 'core/students_list.html', context)


def student_detail(request, registration_no):
    """Student detail view"""
    try:
        student = Student.objects.get(registration_no=registration_no)
        
        # Get related information safely
        try:
            parent_info = ParentInfo.objects.get(registration_no=student)
        except ParentInfo.DoesNotExist:
            parent_info = None
            
        try:
            academic_info = AcademicInfo.objects.get(registration_no=student)
        except AcademicInfo.DoesNotExist:
            academic_info = None
            
        try:
            fee_details = FeeDetails.objects.get(registration_no=student)
        except FeeDetails.DoesNotExist:
            fee_details = None
        
        # Get installments from new FeeInstallments table
        installments = []
        try:
            from core.models import FeeInstallments
            installments = FeeInstallments.objects.filter(registration_no=student).order_by('installment_no')
        except Exception as e:
            print(f"Error fetching installments: {e}")
        
        context = {
            'student': student,
            'parent_info': parent_info,
            'academic_info': academic_info,
            'fee_details': fee_details,
            'installments': installments,
        }
        return render(request, 'core/student_detail.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, f'Student with registration number {registration_no} not found.')
        return redirect('core:students_list')


@login_required
def fee_details(request):
    """Fee details and management"""
    fee_details = FeeDetails.objects.all()
    
    # Filter by payment status
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'pending':
        fee_details = fee_details.filter(fees_remaining__gt=0)
    elif status_filter == 'paid':
        fee_details = fee_details.filter(fees_remaining=0)
    
    context = {
        'fee_details': fee_details,
        'status_filter': status_filter
    }
    return render(request, 'core/fee_details.html', context)


@login_required
def whatsapp_panel(request):
    """WhatsApp reminder panel"""
    # Get students with pending fees - simplified for now
    all_students = Student.objects.all()
    pending_students = []
    
    for student in all_students:
        try:
            fee_details = FeeDetails.objects.get(registration_no=student)
            if fee_details.fees_remaining > 0:
                pending_students.append(student)
        except FeeDetails.DoesNotExist:
            continue
    
    context = {
        'pending_students': pending_students
    }
    return render(request, 'core/whatsapp_panel.html', context)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def send_whatsapp(request):
    """Send WhatsApp messages to selected students"""
    try:
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        message_template = data.get('message', '')
        
        if not student_ids or not message_template:
            return JsonResponse({
                'success': False,
                'message': 'Missing student IDs or message content'
            })
        
        # Initialize Twilio client
        client = Client(settings.TWILIO_SID, settings.TWILIO_AUTH_TOKEN)
        sent_count = 0
        failed_count = 0
        
        for student_id in student_ids:
            try:
                student = Student.objects.get(registration_no=student_id)
                
                # Personalize message
                message = message_template.replace('[Father_Name]', student.parentinfo.father_name)
                message = message.replace('[Student_Name]', student.full_name)
                message = message.replace('[Pending_Amount]', str(student.feedetails.fees_remaining))
                message = message.replace('[Course]', student.academicinfo.enrolled_course)
                
                # Get mobile number (prefer father's mobile)
                mobile = student.parentinfo.father_mobile or student.parentinfo.mother_mobile
                if mobile:
                    # Format mobile number for WhatsApp
                    if not mobile.startswith('+'):
                        mobile = '+91' + mobile  # Assuming Indian numbers
                    
                    # Send message via Twilio
                    message_obj = client.messages.create(
                        body=message,
                        from_=settings.TWILIO_WHATSAPP_NUMBER,
                        to=f'whatsapp:{mobile}'
                    )
                    
                    sent_count += 1
                else:
                    failed_count += 1
                    
            except Exception as e:
                failed_count += 1
                print(f"Failed to send message to student {student_id}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Messages sent to {sent_count} students successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Failed to send messages: {str(e)}'
        })


@login_required
def recent_messages(request):
    """Get recent WhatsApp messages (mock data for now)"""
    # This would typically come from a database table storing message history
    # For now, returning mock data
    messages_data = []
    
    return JsonResponse({
        'success': True,
        'messages': messages_data
    })


@login_required
def reports(request):
    """Reports and analytics page"""
    # Calculate various statistics for reports
    total_students = Student.objects.count()
    
    # Calculate fees properly
    fee_aggregates = FeeDetails.objects.aggregate(
        total_fees=Sum('total_fees'),
        total_remaining=Sum('fees_remaining')
    )
    
    total_revenue = fee_aggregates['total_fees'] or 0
    pending_fees = fee_aggregates['total_remaining'] or 0
    collected_fees = total_revenue - pending_fees
    
    # Course-wise statistics
    course_stats = {}
    students_by_course = Student.objects.all()
    for student in students_by_course:
        course = student.academicinfo.enrolled_course if hasattr(student, 'academicinfo') else 'Unknown'
        if course not in course_stats:
            course_stats[course] = {'count': 0, 'fees_collected': 0, 'fees_pending': 0}
        course_stats[course]['count'] += 1
        if hasattr(student, 'feedetails'):
            course_stats[course]['fees_collected'] += student.feedetails.fees_paid
            course_stats[course]['fees_pending'] += student.feedetails.fees_remaining
    
    context = {
        'total_students': total_students,
        'total_revenue': total_revenue,
        'collected_fees': collected_fees,
        'pending_fees': pending_fees,
        'course_stats': course_stats,
    }
    return render(request, 'core/reports.html', context)


def login_view(request):
    """Login view"""
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, 'Login successful!')
            return redirect('core:dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'core/login.html')


@login_required
def logout_view(request):
    """Logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('core:login')


@login_required
def student_edit(request, registration_no):
    """Edit student information"""
    print(f"=== STUDENT EDIT VIEW CALLED ===")
    print(f"Method: {request.method}")
    print(f"Registration No: {registration_no}")
    
    try:
        student = Student.objects.get(registration_no=registration_no)
        print(f"Student found: {student.full_name}")
        
        # Get related information safely
        try:
            parent_info = ParentInfo.objects.get(registration_no=student)
        except ParentInfo.DoesNotExist:
            parent_info = None
            
        try:
            academic_info = AcademicInfo.objects.get(registration_no=student)
        except AcademicInfo.DoesNotExist:
            academic_info = None
            
        try:
            fee_details = FeeDetails.objects.get(registration_no=student)
        except FeeDetails.DoesNotExist:
            fee_details = None
        
        if request.method == 'POST':
            print("=== POST REQUEST RECEIVED ===")
            print("POST Data:", dict(request.POST))
            
            try:
                # Update student information
                student_form = StudentRegistrationForm(request.POST, instance=student)
                parent_form = ParentInfoForm(request.POST, instance=parent_info) if parent_info else ParentInfoForm(request.POST)
                academic_form = AcademicInfoForm(request.POST, instance=academic_info) if academic_info else AcademicInfoForm(request.POST)
                fee_form = FeeDetailsForm(request.POST, instance=fee_details) if fee_details else FeeDetailsForm(request.POST)
                
                # Check form validation
                forms_valid = {
                    'student': student_form.is_valid(),
                    'parent': parent_form.is_valid(), 
                    'academic': academic_form.is_valid(),
                    'fee': fee_form.is_valid()
                }
                print("Form validation results:", forms_valid)
                
                # Print detailed validation errors immediately
                if not forms_valid['student']:
                    print("Student form errors:", dict(student_form.errors))
                if not forms_valid['parent']:
                    print("Parent form errors:", dict(parent_form.errors))
                if not forms_valid['academic']:
                    print("Academic form errors:", dict(academic_form.errors))
                if not forms_valid['fee']:
                    print("Fee form errors:", dict(fee_form.errors))
                
                if all(forms_valid.values()):
                    print("All forms valid, saving...")
                    # Save student
                    student = student_form.save()
                    print("Student saved:", student)
                    
                    # Save parent info
                    parent_obj = parent_form.save(commit=False)
                    parent_obj.registration_no = student
                    parent_obj.save()
                    
                    # Save academic info
                    academic_obj = academic_form.save(commit=False)
                    academic_obj.registration_no = student
                    academic_obj.save()
                    
                    # Save fee details with enhanced validation
                    fee_obj = fee_form.save(commit=False)
                    fee_obj.registration_no = student
                    
                    # Enhanced validation and safety checks
                    try:
                        # Validate the fee object before saving
                        fee_obj.full_clean()
                        
                        # Safety check: Ensure fees_remaining is not incorrectly set to 0
                        if fee_obj.fees_remaining <= 0 and fee_obj.total_fees > 0:
                            # If fees_remaining is 0 but total_fees exists, assume no payments made
                            fee_obj.fees_remaining = fee_obj.total_fees
                        
                        fee_obj.save()
                        
                    except ValidationError as e:
                        print("Fee details validation error:", e.message_dict)
                        error_messages = []
                        for field, errors in e.message_dict.items():
                            for error in errors:
                                error_messages.append(f"Fee {field}: {error}")
                        
                        messages.error(request, f'Fee details validation failed: {"; ".join(error_messages)}')
                        # Re-render the form with errors
                        context = {
                            'student': student,
                            'student_form': student_form,
                            'parent_form': parent_form,
                            'academic_form': academic_form,
                            'fee_form': fee_form,
                            'branches': Branch.objects.all(),
                            'is_edit': True
                        }
                        return render(request, 'core/student_edit.html', context)
                    
                    # Handle dynamic installment updates
                    try:
                        from core.models import FeeInstallments
                        
                        # Delete existing installments for this student
                        FeeInstallments.objects.filter(registration_no=student).delete()
                        
                        # Create new installments based on form data
                        num_installments = int(request.POST.get('number_of_installments', 1))
                        if num_installments > 1:
                            for i in range(1, num_installments + 1):
                                amount = None
                                due_date = None
                                
                                # Handle first 3 installments (static fields)
                                if i == 1:
                                    amount = request.POST.get('first_installment')
                                    due_date = request.POST.get('first_installment_date')
                                elif i == 2:
                                    amount = request.POST.get('second_installment')
                                    due_date = request.POST.get('second_installment_date')
                                elif i == 3:
                                    amount = request.POST.get('third_installment')
                                    due_date = request.POST.get('third_installment_date')
                                else:
                                    # Handle dynamic installments (4+)
                                    amount = request.POST.get(f'installment_{i}_amount')
                                    due_date = request.POST.get(f'installment_{i}_date')
                                
                                if amount and due_date:
                                    try:
                                        # Create installment with validation
                                        installment = FeeInstallments(
                                            registration_no=student,
                                            installment_no=i,
                                            amount=float(amount),
                                            due_date=due_date,
                                            status='Due'
                                        )
                                        # Validate before saving
                                        installment.full_clean()
                                        installment.save()
                                        
                                    except ValidationError as e:
                                        print(f"Installment {i} validation error:", e.message_dict)
                                        messages.warning(request, f'Warning: Installment {i} had validation issues and was skipped')
                                    except ValueError as e:
                                        print(f"Installment {i} value error:", str(e))
                                        messages.warning(request, f'Warning: Installment {i} had invalid amount and was skipped')
                    except Exception as installment_error:
                        print(f"Warning: Could not update installments: {installment_error}")
                    
                    messages.success(request, f'Student {student.full_name} updated successfully!')
                    return redirect('core:student_detail', registration_no=student.registration_no)
                else:
                    print("Form validation failed!")
                    # Collect all form errors
                    all_errors = []
                    form_names = ['student', 'parent', 'academic', 'fee']
                    forms = [student_form, parent_form, academic_form, fee_form]
                    
                    for form_name, form in zip(form_names, forms):
                        if form.errors:
                            print(f"{form_name} form errors:", dict(form.errors))
                            for field, errors in form.errors.items():
                                for error in errors:
                                    all_errors.append(f"{form_name}.{field}: {error}")
                    
                    print("All errors:", all_errors)
                    messages.error(request, f'Please correct the following errors: {", ".join(all_errors)}')
                    
            except Exception as e:
                messages.error(request, f'Error updating student: {str(e)}')
        else:
            # Initialize forms with existing data
            student_form = StudentRegistrationForm(instance=student)
            parent_form = ParentInfoForm(instance=parent_info) if parent_info else ParentInfoForm()
            academic_form = AcademicInfoForm(instance=academic_info) if academic_info else AcademicInfoForm()
            fee_form = FeeDetailsForm(instance=fee_details) if fee_details else FeeDetailsForm()
        
        # Get all branches for the form
        branches = Branch.objects.all()
        
        context = {
            'student': student,
            'student_form': student_form,
            'parent_form': parent_form,
            'academic_form': academic_form,
            'fee_form': fee_form,
            'branches': branches,
            'is_edit': True
        }
        return render(request, 'core/student_edit.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, f'Student with registration number {registration_no} not found.')
        return redirect('core:students_list')


@login_required
def fee_payment(request, registration_no):
    """Record fee payment for a student"""
    try:
        student = Student.objects.get(registration_no=registration_no)
        fee_details = FeeDetails.objects.get(registration_no=student)
        
        if request.method == 'POST':
            form = FeePaymentForm(request.POST)
            if form.is_valid():
                payment_amount = form.cleaned_data['payment_amount']
                
                # Validate payment amount
                if payment_amount > fee_details.fees_remaining:
                    messages.error(request, f'Payment amount (₹{payment_amount}) cannot exceed remaining fees (₹{fee_details.fees_remaining})')
                else:
                    # Update fee details
                    fee_details.fees_remaining -= payment_amount
                    fee_details.save()
                    
                    messages.success(request, f'Payment of ₹{payment_amount} recorded successfully! Remaining fees: ₹{fee_details.fees_remaining}')
                    return redirect('core:student_detail', registration_no=student.registration_no)
            else:
                messages.error(request, 'Please correct the errors in the form.')
        else:
            form = FeePaymentForm()
        
        context = {
            'student': student,
            'fee_details': fee_details,
            'form': form
        }
        return render(request, 'core/fee_payment.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, f'Student with registration number {registration_no} not found.')
        return redirect('core:students_list')
    except FeeDetails.DoesNotExist:
        messages.error(request, f'Fee details not found for student {registration_no}.')
        return redirect('core:student_detail', registration_no=registration_no)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def quick_fee_update(request):
    """Quick fee update via AJAX"""
    try:
        data = json.loads(request.body)
        registration_no = data.get('registration_no')
        payment_amount = float(data.get('payment_amount', 0))
        
        if not registration_no or payment_amount <= 0:
            return JsonResponse({
                'success': False,
                'message': 'Invalid registration number or payment amount'
            })
        
        student = Student.objects.get(registration_no=registration_no)
        fee_details = FeeDetails.objects.get(registration_no=student)
        
        if payment_amount > fee_details.fees_remaining:
            return JsonResponse({
                'success': False,
                'message': f'Payment amount (₹{payment_amount}) cannot exceed remaining fees (₹{fee_details.fees_remaining})'
            })
        
        # Update fee details
        fee_details.fees_remaining -= payment_amount
        fee_details.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Payment of ₹{payment_amount} recorded successfully!',
            'new_remaining': float(fee_details.fees_remaining),
            'fees_paid': float(fee_details.fees_paid)
        })
        
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Student not found'
        })
    except FeeDetails.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Fee details not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing payment: {str(e)}'
        })


# Reports and Analytics Views


@login_required
def analytics_dashboard(request):
    """Analytics dashboard with comprehensive statistics and charts - Updated for FeeInstallments schema"""
    from core.models import FeeInstallments
    from datetime import date
    
    # Basic statistics
    total_students = Student.objects.count()
    
    # Calculate fees from both FeeDetails and FeeInstallments
    fee_aggregates = FeeDetails.objects.aggregate(
        total_fees=Sum('total_fees'),
        total_remaining=Sum('fees_remaining')
    )
    
    total_fees = fee_aggregates['total_fees'] or 0
    total_pending_fees = fee_aggregates['total_remaining'] or 0
    total_fees_collected = total_fees - total_pending_fees
    
    # NEW: Installment-based statistics
    installment_stats = FeeInstallments.objects.aggregate(
        total_installments=models.Count('installment_id'),
        total_installment_amount=Sum('amount'),
        paid_installments=models.Count('installment_id', filter=models.Q(status='Paid')),
        paid_amount=Sum('amount', filter=models.Q(status='Paid')),
        overdue_installments=models.Count('installment_id', filter=models.Q(due_date__lt=date.today(), status__in=['Due', 'Pending'])),
        due_today=models.Count('installment_id', filter=models.Q(due_date=date.today(), status__in=['Due', 'Pending']))
    )
    
    # NEW: Monthly installment trends
    from django.db.models import Extract
    monthly_installments = FeeInstallments.objects.filter(
        due_date__year=date.today().year
    ).annotate(
        month=Extract('due_date', 'month')
    ).values('month').annotate(
        total_amount=Sum('amount'),
        paid_amount=Sum('amount', filter=models.Q(status='Paid')),
        installment_count=models.Count('installment_id')
    ).order_by('month')
    
    # Branch-wise statistics (enhanced with installments)
    branch_stats = Branch.objects.annotate(
        student_count=models.Count('academicinfo__registration_no'),
        total_fees=Sum('academicinfo__registration_no__feedetails__total_fees'),
        collected_fees=Sum('academicinfo__registration_no__feedetails__total_fees') - 
                      Sum('academicinfo__registration_no__feedetails__fees_remaining'),
        pending_fees=Sum('academicinfo__registration_no__feedetails__fees_remaining'),
        # NEW: Installment stats per branch
        total_installments=models.Count('academicinfo__registration_no__feeinstallments'),
        paid_installments=models.Count('academicinfo__registration_no__feeinstallments', 
                                     filter=models.Q(academicinfo__registration_no__feeinstallments__status='Paid'))
    ).values('branch_name', 'student_count', 'total_fees', 'collected_fees', 'pending_fees', 
             'total_installments', 'paid_installments')
    
    # Course-wise statistics (enhanced with installments)
    course_stats = AcademicInfo.objects.values('enrolled_course').annotate(
        student_count=models.Count('registration_no'),
        total_fees=Sum('registration_no__feedetails__total_fees'),
        collected_fees=Sum('registration_no__feedetails__total_fees') - 
                      Sum('registration_no__feedetails__fees_remaining'),
        pending_fees=Sum('registration_no__feedetails__fees_remaining'),
        # NEW: Installment stats per course
        total_installments=models.Count('registration_no__feeinstallments'),
        paid_installments=models.Count('registration_no__feeinstallments', 
                                     filter=models.Q(registration_no__feeinstallments__status='Paid'))
    )
    
    # Payment status distribution (enhanced)
    fully_paid = FeeDetails.objects.filter(fees_remaining=0).count()
    partially_paid = FeeDetails.objects.filter(fees_remaining__gt=0, fees_remaining__lt=models.F('total_fees')).count()
    not_paid = FeeDetails.objects.filter(fees_remaining=models.F('total_fees')).count()
    
    # NEW: Installment status distribution
    installment_status_stats = FeeInstallments.objects.values('status').annotate(
        count=models.Count('installment_id'),
        total_amount=Sum('amount')
    )
    
    # NEW: Upcoming installments (next 30 days)
    from datetime import timedelta
    upcoming_installments = FeeInstallments.objects.filter(
        due_date__range=[date.today(), date.today() + timedelta(days=30)],
        status__in=['Due', 'Pending']
    ).select_related('registration_no').order_by('due_date')[:10]
    
    context = {
        'total_students': total_students,
        'total_fees_collected': total_fees_collected,
        'total_pending_fees': total_pending_fees,
        'branch_stats': list(branch_stats),
        'course_stats': list(course_stats),
        'payment_distribution': {
            'fully_paid': fully_paid,
            'partially_paid': partially_paid,
            'not_paid': not_paid
        },
        # NEW: Installment-based analytics
        'installment_stats': installment_stats,
        'monthly_installments': list(monthly_installments),
        'installment_status_stats': list(installment_status_stats),
        'upcoming_installments': upcoming_installments,
    }
    
    return render(request, 'core/analytics_dashboard.html', context)


@login_required
def generate_fees_report_pdf(request):
    """Generate comprehensive fees report as PDF - Updated for FeeInstallments schema"""
    from core.models import FeeInstallments
    from datetime import date
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="fees_report_enhanced.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph("Enhanced Fees Management Report", title_style))
    story.append(Spacer(1, 12))
    
    # Summary statistics
    total_students = Student.objects.count()
    
    # Calculate fees properly
    fee_aggregates = FeeDetails.objects.aggregate(
        total_fees=Sum('total_fees'),
        total_remaining=Sum('fees_remaining')
    )
    
    total_fees = fee_aggregates['total_fees'] or 0
    total_pending_fees = fee_aggregates['total_remaining'] or 0
    total_fees_collected = total_fees - total_pending_fees
    
    # NEW: Installment statistics
    installment_stats = FeeInstallments.objects.aggregate(
        total_installments=models.Count('installment_id'),
        total_installment_amount=Sum('amount'),
        paid_installments=models.Count('installment_id', filter=models.Q(status='Paid')),
        paid_amount=Sum('amount', filter=models.Q(status='Paid')),
        overdue_installments=models.Count('installment_id', filter=models.Q(due_date__lt=date.today(), status__in=['Due', 'Pending']))
    )
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Students', str(total_students)],
        ['Total Fees Collected', f'₹{total_fees_collected:,.2f}'],
        ['Total Pending Fees', f'₹{total_pending_fees:,.2f}'],
        ['Collection Rate', f'{(total_fees_collected/(total_fees_collected+total_pending_fees)*100):.1f}%' if (total_fees_collected+total_pending_fees) > 0 else '0%'],
        ['', ''],  # Separator
        ['INSTALLMENT ANALYTICS', ''],
        ['Total Installments', str(installment_stats['total_installments'] or 0)],
        ['Paid Installments', str(installment_stats['paid_installments'] or 0)],
        ['Overdue Installments', str(installment_stats['overdue_installments'] or 0)],
        ['Installment Amount Paid', f'₹{installment_stats["paid_amount"] or 0:,.2f}'],
        ['Installment Payment Rate', f'{((installment_stats["paid_installments"] or 0)/(installment_stats["total_installments"] or 1)*100):.1f}%']
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # NEW: Installment Status Breakdown
    story.append(Paragraph("Installment Status Breakdown", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    installment_status_data = [['Status', 'Count', 'Total Amount']]
    installment_status_stats = FeeInstallments.objects.values('status').annotate(
        count=models.Count('installment_id'),
        total_amount=Sum('amount')
    )
    
    for status_stat in installment_status_stats:
        installment_status_data.append([
            status_stat['status'] or 'Unknown',
            str(status_stat['count'] or 0),
            f'₹{status_stat["total_amount"] or 0:,.2f}'
        ])
    
    installment_status_table = Table(installment_status_data)
    installment_status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(installment_status_table)
    story.append(Spacer(1, 20))
    
    # Student-wise fee details with installment info
    story.append(Paragraph("Student-wise Fee & Installment Details", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    students_data = [['Reg No', 'Name', 'Course', 'Total Fees', 'Paid', 'Pending', 'Installments']]
    
    for student in Student.objects.select_related('feedetails', 'academicinfo').all()[:20]:  # Limit to first 20
        fee_details = getattr(student, 'feedetails', None)
        academic_info = getattr(student, 'academicinfo', None)
        
        # Get installment count for this student
        installment_count = FeeInstallments.objects.filter(registration_no=student).count()
        
        if fee_details:
            students_data.append([
                str(student.registration_no),
                student.full_name[:15],  # Truncate long names
                academic_info.enrolled_course[:12] if academic_info else 'N/A',
                f'₹{fee_details.total_fees:,.0f}',
                f'₹{fee_details.fees_paid:,.0f}',
                f'₹{fee_details.fees_remaining:,.0f}',
                str(installment_count)
            ])
    
    students_table = Table(students_data)
    students_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(students_table)
    
    doc.build(story)
    return response


@login_required
def generate_analytics_pdf(request):
    """Generate analytics report with charts as PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="analytics_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph("Analytics Report", title_style))
    story.append(Spacer(1, 12))
    
    # Branch-wise analysis (enhanced with installments)
    branch_stats = Branch.objects.annotate(
        student_count=models.Count('academicinfo__registration_no'),
        collected_fees=Sum('academicinfo__registration_no__feedetails__total_fees') - 
                      Sum('academicinfo__registration_no__feedetails__fees_remaining'),
        total_installments=models.Count('academicinfo__registration_no__feeinstallments'),
        paid_installments=models.Count('academicinfo__registration_no__feeinstallments', 
                                     filter=models.Q(academicinfo__registration_no__feeinstallments__status='Paid'))
    ).values('branch_name', 'student_count', 'collected_fees', 'total_installments', 'paid_installments')
    
    story.append(Paragraph("Branch-wise Statistics", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    branch_data = [['Branch', 'Students', 'Fees Collected', 'Total Installments', 'Paid Installments']]
    for branch in branch_stats:
        branch_data.append([
            branch['branch_name'] or 'N/A',
            str(branch['student_count'] or 0),
            f"₹{branch['collected_fees'] or 0:,.0f}",
            str(branch['total_installments'] or 0),
            str(branch['paid_installments'] or 0)
        ])
    
    branch_table = Table(branch_data)
    branch_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(branch_table)
    
    doc.build(story)
    return response


@login_required
def export_data_excel(request):
    """Export student and fee data to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_fees_data.xlsx"'
    
    # Create DataFrame with student and fee data
    students_data = []
    for student in Student.objects.select_related('feedetails', 'academicinfo', 'parentinfo').all():
        fee_details = getattr(student, 'feedetails', None)
        academic_info = getattr(student, 'academicinfo', None)
        parent_info = getattr(student, 'parentinfo', None)
        
        students_data.append({
            'Registration No': student.registration_no,
            'Full Name': student.full_name,
            'Gender': student.get_gender_display(),
            'Date of Birth': student.date_of_birth,
            'Date of Admission': student.date_of_admission,
            'Course': academic_info.enrolled_course if academic_info else 'N/A',
            'Branch': academic_info.branch_code.branch_name if academic_info and academic_info.branch_code else 'N/A',
            'Father Name': parent_info.father_name if parent_info else 'N/A',
            'Mother Name': parent_info.mother_name if parent_info else 'N/A',
            'Father Mobile': parent_info.father_mobile if parent_info else 'N/A',
            'Total Fees': fee_details.total_fees if fee_details else 0,
            'Fees Paid': fee_details.fees_paid if fee_details else 0,
            'Fees Remaining': fee_details.fees_remaining if fee_details else 0,
            'Installments': fee_details.number_of_installments if fee_details else 0,
            'Fees per Installment': fee_details.fees_per_installment if fee_details else 0
        })
    
    df = pd.DataFrame(students_data)
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Student Data', index=False)
        
        # Add summary sheet
        summary_data = {
            'Metric': ['Total Students', 'Total Fees Collected', 'Total Pending Fees', 'Average Fee per Student'],
            'Value': [
                len(students_data),
                sum(row['Fees Paid'] for row in students_data),
                sum(row['Fees Remaining'] for row in students_data),
                sum(row['Total Fees'] for row in students_data) / len(students_data) if students_data else 0
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    return response


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def remove_student(request):
    """Remove a student and all related data"""
    try:
        data = json.loads(request.body)
        registration_no = data.get('registration_no')
        
        if not registration_no:
            return JsonResponse({
                'success': False,
                'message': 'Registration number is required'
            })
        
        student = Student.objects.get(registration_no=registration_no)
        student_name = student.full_name
        
        # Delete the student (cascade will handle related records)
        student.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Student {student_name} has been removed successfully'
        })
        
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Student not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error removing student: {str(e)}'
        })


@login_required
@csrf_exempt
def send_bulk_whatsapp(request):
    """Send WhatsApp messages to multiple selected students"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        data = json.loads(request.body)
        students = data.get('students', [])
        
        if not students:
            return JsonResponse({'success': False, 'message': 'No students selected'})
        
        sent_count = 0
        failed_count = 0
        
        for student in students:
            try:
                # Get full student data from database
                student_obj = Student.objects.select_related('feedetails', 'parentinfo').get(
                    registration_no=student['registration_no']
                )
                
                # Prepare message
                message = f"""Dear {student_obj.first_name} {student_obj.last_name},

This is a reminder that your payment of ₹{student_obj.feedetails.fees_remaining} for your course is due soon. Please pay at the earliest.

Thank you,
Pillay Sir's ICSE Classes"""
                
                # Here you would integrate with your WhatsApp service
                # For now, we'll just log the attempt
                print(f"Would send WhatsApp to {student['name']} at {student['mobile']}")
                sent_count += 1
                
            except Student.DoesNotExist:
                print(f"Student {student['registration_no']} not found")
                failed_count += 1
            except Exception as e:
                print(f"Error sending to {student['name']}: {e}")
                failed_count += 1
        
        return JsonResponse({
            'success': True,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Messages processed: {sent_count} sent, {failed_count} failed'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing bulk WhatsApp: {str(e)}'
        })


@login_required
def manual_whatsapp_reminders(request):
    """Manually trigger WhatsApp reminders"""
    if request.method == 'POST':
        try:
            # Get the path to the WhatsApp reminder script
            script_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'send_whatsapp_reminders.py')
            
            # Run the script with proper encoding and Python executable
            import sys
            python_executable = sys.executable
            
            # Set environment for UTF-8 encoding
            env = os.environ.copy()
            env['PYTHONIOENCODING'] = 'utf-8'
            
            result = subprocess.run([python_executable, script_path], 
                                  capture_output=True, 
                                  text=True, 
                                  encoding='utf-8',
                                  env=env)
            
            if result.returncode == 0:
                return JsonResponse({
                    'success': True,
                    'message': 'WhatsApp reminders sent successfully!',
                    'output': result.stdout
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Error running WhatsApp reminder script',
                    'error': result.stderr
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error triggering WhatsApp reminders: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def blacklisted_students(request):
    """View for blacklisted students (overdue fees)"""
    from datetime import date
    
    # Get students with overdue fees
    blacklisted = []
    students_with_fees = Student.objects.select_related('feedetails', 'academicinfo__branch_code').filter(
        feedetails__fees_remaining__gt=0
    )
    
    for student in students_with_fees:
        fee_details = student.feedetails
        if fee_details.is_overdue:
            blacklisted.append({
                'student': student,
                'fee_details': fee_details,
                'academic_info': student.academicinfo if hasattr(student, 'academicinfo') else None,
                'days_overdue': fee_details.days_overdue,
                'next_due_date': fee_details.next_due_date
            })
    
    # Sort by days overdue (most overdue first)
    blacklisted.sort(key=lambda x: x['days_overdue'], reverse=True)
    
    context = {
        'blacklisted_students': blacklisted,
        'total_blacklisted': len(blacklisted)
    }
    return render(request, 'core/blacklisted_students.html', context)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def mark_as_paid(request):
    """Mark a student's fees as fully paid"""
    try:
        data = json.loads(request.body)
        registration_no = data.get('registration_no')
        
        if not registration_no:
            return JsonResponse({
                'success': False,
                'message': 'Registration number is required'
            })
        
        student = Student.objects.get(registration_no=registration_no)
        fee_details = FeeDetails.objects.get(registration_no=student)
        
        # Mark as paid
        fee_details.fees_remaining = 0
        fee_details.save()
        
        return JsonResponse({
            'success': True,
            'message': f'{student.full_name} has been marked as paid',
            'student_name': student.full_name
        })
        
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Student not found'
        })
    except FeeDetails.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Fee details not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error marking as paid: {str(e)}'
        })


@login_required
def custom_whatsapp_panel(request):
    """Enhanced WhatsApp panel with custom messaging and file upload"""
    all_students = Student.objects.select_related('feedetails', 'parentinfo', 'academicinfo').all()
    
    # Get recent custom messages
    recent_messages = CustomMessageLog.objects.select_related('student').order_by('-timestamp')[:10]
    
    context = {
        'all_students': all_students,
        'recent_messages': recent_messages
    }
    return render(request, 'core/custom_whatsapp_panel.html', context)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def send_custom_whatsapp(request):
    """Send custom WhatsApp messages with optional file attachment"""
    try:
        # Handle both form data and JSON
        if request.content_type and request.content_type.startswith('multipart/form-data'):
            message_text = request.POST.get('message_text', '')
            recipient_type = request.POST.get('recipient_type', 'selected')
            selected_students = request.POST.getlist('selected_students[]')
            attachment = request.FILES.get('attachment')
        else:
            data = json.loads(request.body)
            message_text = data.get('message_text', '')
            recipient_type = data.get('recipient_type', 'selected')
            selected_students = data.get('selected_students', [])
            attachment = None
        
        if not message_text:
            return JsonResponse({
                'success': False,
                'message': 'Message text is required'
            })
        
        # Determine recipients
        if recipient_type == 'all':
            students = Student.objects.select_related('parentinfo').all()
        else:
            if not selected_students:
                return JsonResponse({
                    'success': False,
                    'message': 'No students selected'
                })
            students = Student.objects.select_related('parentinfo').filter(
                registration_no__in=selected_students
            )
        
        # Handle file attachment
        attachment_path = None
        if attachment:
            # Save the uploaded file
            import os
            from django.conf import settings
            
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'whatsapp_attachments')
            os.makedirs(upload_dir, exist_ok=True)
            
            attachment_path = os.path.join(upload_dir, attachment.name)
            with open(attachment_path, 'wb+') as destination:
                for chunk in attachment.chunks():
                    destination.write(chunk)
        
        sent_count = 0
        failed_count = 0
        
        for student in students:
            try:
                # Log the message
                CustomMessageLog.objects.create(
                    student=student,
                    message_text=message_text,
                    attachment=attachment_path if attachment_path else None,
                    sent_status='SUCCESS'
                )
                
                # Here you would integrate with your WhatsApp service
                # For now, we'll simulate sending
                print(f"Custom message sent to {student.full_name}")
                sent_count += 1
                
            except Exception as e:
                CustomMessageLog.objects.create(
                    student=student,
                    message_text=message_text,
                    attachment=attachment_path if attachment_path else None,
                    sent_status='FAILED'
                )
                print(f"Failed to send to {student.full_name}: {e}")
                failed_count += 1
        
        # Send summary message to admin
        try:
            admin_summary = f"""📱 Custom Message Summary - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📝 MESSAGE SENT:
"{message_text[:100]}{'...' if len(message_text) > 100 else ''}"

👥 RECIPIENTS ({sent_count + failed_count} total):
✅ Successfully sent: {sent_count}
❌ Failed: {failed_count}

📋 STUDENT LIST:"""
            
            # Add student names to summary
            for student in students:
                status = "✅" if CustomMessageLog.objects.filter(
                    student=student, 
                    message_text=message_text,
                    sent_status='SUCCESS'
                ).exists() else "❌"
                admin_summary += f"\n{status} {student.full_name} ({student.registration_no})"
            
            admin_summary += f"\n\nPillay Sir's ICSE Classes - Custom Message System"
            
            # Here you would send this summary to your WhatsApp number
            # For now, we'll log it
            print("Admin Summary:", admin_summary)
            
            # Also create Excel log for custom messages
            try:
                import pandas as pd
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                
                # Create Excel file for custom messages
                excel_file = f'logs/custom_messages_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
                # Prepare data for Excel
                excel_data = []
                for student in students:
                    status = "Sent" if CustomMessageLog.objects.filter(
                        student=student, 
                        message_text=message_text,
                        sent_status='SUCCESS'
                    ).exists() else "Failed"
                    
                    excel_data.append({
                        'Student Name': student.full_name,
                        'Registration No': student.registration_no,
                        'Course': student.academicinfo.enrolled_course if hasattr(student, 'academicinfo') else 'N/A',
                        'Father Mobile': student.parentinfo.father_mobile if hasattr(student, 'parentinfo') else 'N/A',
                        'Message': message_text[:100] + '...' if len(message_text) > 100 else message_text,
                        'Status': status,
                        'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                
                # Create DataFrame and save to Excel
                df = pd.DataFrame(excel_data)
                
                # Ensure logs directory exists
                import os
                os.makedirs('logs', exist_ok=True)
                
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Custom Messages', index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Custom Messages']
                    
                    # Style the header row
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                
                print(f"📊 Custom message Excel log created: {excel_file}")
                
            except Exception as excel_error:
                print(f"⚠️ Could not create Excel log: {excel_error}")
            
        except Exception as e:
            print(f"Failed to prepare admin summary: {e}")
        
        return JsonResponse({
            'success': True,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Custom messages processed: {sent_count} sent, {failed_count} failed'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error sending custom messages: {str(e)}'
        })


@login_required
def installment_management(request):
    """Installment Management Panel"""
    # Get all installments with student details
    installments = FeeInstallments.objects.select_related(
        'registration_no', 
        'registration_no__academicinfo', 
        'registration_no__parentinfo'
    ).order_by('due_date', 'registration_no')
    
    # Filter by status if requested
    status_filter = request.GET.get('status')
    if status_filter:
        installments = installments.filter(status=status_filter)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        installments = installments.filter(
            Q(registration_no__first_name__icontains=search_query) |
            Q(registration_no__last_name__icontains=search_query) |
            Q(registration_no__registration_no__icontains=search_query)
        )
    
    # Update overdue statuses
    FeeInstallments.update_overdue_statuses()
    
    # Statistics
    stats = {
        'total_installments': FeeInstallments.objects.count(),
        'due_count': FeeInstallments.objects.filter(status='Due').count(),
        'pending_count': FeeInstallments.objects.filter(status='Pending').count(),
        'paid_count': FeeInstallments.objects.filter(status='Paid').count(),
        'total_due_amount': FeeInstallments.objects.filter(status__in=['Due', 'Pending']).aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_paid_amount': FeeInstallments.objects.filter(status='Paid').aggregate(Sum('amount'))['amount__sum'] or 0,
    }
    
    context = {
        'installments': installments,
        'stats': stats,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    
    return render(request, 'core/installment_management.html', context)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def mark_installment_paid(request):
    """Mark an installment as paid via AJAX"""
    try:
        data = json.loads(request.body)
        installment_id = data.get('installment_id')
        
        installment = get_object_or_404(FeeInstallments, installment_id=installment_id)
        
        if installment.status == 'Paid':
            return JsonResponse({
                'success': False,
                'message': 'Installment is already marked as paid'
            })
        
        # Mark as paid
        installment.mark_as_paid()
        
        # Get updated student fee details
        try:
            fee_details = installment.registration_no.feedetails
            remaining_fees = fee_details.fees_remaining
        except FeeDetails.DoesNotExist:
            remaining_fees = 0
        
        return JsonResponse({
            'success': True,
            'message': f'Installment {installment.installment_no} marked as paid successfully!',
            'paid_date': installment.paid_date.strftime('%Y-%m-%d'),
            'remaining_fees': float(remaining_fees)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error marking installment as paid: {str(e)}'
        })


@login_required
def blacklisted_students_new(request):
    """Enhanced Blacklisted Students Panel based on installment status"""
    # Students with pending installments
    blacklisted_students = Student.objects.filter(
        installments__status='Pending'
    ).distinct().select_related('academicinfo', 'parentinfo', 'feedetails')
    
    # Add installment details for each student
    student_data = []
    for student in blacklisted_students:
        pending_installments = student.installments.filter(status='Pending').order_by('due_date')
        overdue_installments = student.installments.filter(status='Due').order_by('due_date')
        
        # Calculate total overdue amount
        total_overdue = pending_installments.aggregate(Sum('amount'))['amount__sum'] or 0
        total_overdue += overdue_installments.filter(due_date__lt=datetime.now().date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        student_data.append({
            'student': student,
            'pending_installments': pending_installments,
            'overdue_installments': overdue_installments,
            'total_overdue_amount': total_overdue,
            'days_overdue': max([inst.days_overdue for inst in pending_installments] + [0])
        })
    
    context = {
        'student_data': student_data,
        'total_blacklisted': len(student_data)
    }
    
    return render(request, 'core/blacklisted_students_new.html', context)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_installments_from_fee_details(request):
    """Create installments from existing fee details"""
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        
        student = get_object_or_404(Student, registration_no=student_id)
        
        # Check if installments already exist
        if student.installments.exists():
            return JsonResponse({
                'success': False,
                'message': 'Installments already exist for this student'
            })
        
        # Get fee details
        try:
            fee_details = student.feedetails
        except FeeDetails.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'No fee details found for this student'
            })
        
        # Create installments based on fee details
        installments_created = 0
        
        # First installment
        if fee_details.first_installment and fee_details.first_installment_date:
            FeeInstallments.objects.create(
                registration_no=student,
                installment_no=1,
                amount=fee_details.first_installment,
                due_date=fee_details.first_installment_date
            )
            installments_created += 1
        
        # Second installment
        if fee_details.second_installment and fee_details.second_installment_date:
            FeeInstallments.objects.create(
                registration_no=student,
                installment_no=2,
                amount=fee_details.second_installment,
                due_date=fee_details.second_installment_date
            )
            installments_created += 1
        
        # Third installment
        if fee_details.third_installment and fee_details.third_installment_date:
            FeeInstallments.objects.create(
                registration_no=student,
                installment_no=3,
                amount=fee_details.third_installment,
                due_date=fee_details.third_installment_date
            )
            installments_created += 1
        
        # If no specific installments, create equal installments
        if installments_created == 0 and fee_details.number_of_installments > 0:
            installment_amount = fee_details.fees_per_installment
            from datetime import date, timedelta
            
            for i in range(fee_details.number_of_installments):
                due_date = date.today() + timedelta(days=30 * (i + 1))  # Monthly installments
                FeeInstallments.objects.create(
                    registration_no=student,
                    installment_no=i + 1,
                    amount=installment_amount,
                    due_date=due_date
                )
                installments_created += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{installments_created} installments created successfully',
            'installments_created': installments_created
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating installments: {str(e)}'
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_overdue_statuses(request):
    """Manually trigger overdue status updates"""
    try:
        updated_count = FeeInstallments.update_overdue_statuses()
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} installments updated to Pending status',
            'updated_count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating overdue statuses: {str(e)}'
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def run_installment_reminders(request):
    """Run the installment-based WhatsApp reminder script"""
    try:
        import subprocess
        import os
        from datetime import datetime
        
        # Path to the installment reminder script
        script_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'send_whatsapp_reminders_installments.py')
        
        if not os.path.exists(script_path):
            return JsonResponse({
                'success': False,
                'message': 'Installment reminder script not found'
            })
        
        # Get installments that need reminders today
        from core.models import FeeInstallments
        from datetime import date
        
        today = date.today()
        
        # Get all non-paid installments
        all_installments = FeeInstallments.objects.exclude(status='Paid').select_related('registration_no')
        
        reminder_installments = []
        for installment in all_installments:
            days_difference = (today - installment.due_date).days
            if days_difference in [-3, 1, 4, 7, 10]:
                reminder_installments.append(installment)
        
        if not reminder_installments:
            return JsonResponse({
                'success': True,
                'message': 'No installments require reminders today.',
                'reminders_count': 0
            })
        
        # Count reminders by type
        reminder_types = {}
        for installment in reminder_installments:
            days_diff = (today - installment.due_date).days
            if days_diff == -3:
                reminder_types['reminder_1'] = reminder_types.get('reminder_1', 0) + 1
            elif days_diff == 1:
                reminder_types['reminder_2'] = reminder_types.get('reminder_2', 0) + 1
            elif days_diff == 4:
                reminder_types['reminder_3'] = reminder_types.get('reminder_3', 0) + 1
            elif days_diff == 7:
                reminder_types['last_reminder'] = reminder_types.get('last_reminder', 0) + 1
            elif days_diff == 10:
                reminder_types['discontinuation'] = reminder_types.get('discontinuation', 0) + 1
        
        # Create summary message
        summary_parts = []
        if 'reminder_1' in reminder_types:
            summary_parts.append(f"{reminder_types['reminder_1']} Reminder-1 (3 days before)")
        if 'reminder_2' in reminder_types:
            summary_parts.append(f"{reminder_types['reminder_2']} Reminder-2 (1 day after)")
        if 'reminder_3' in reminder_types:
            summary_parts.append(f"{reminder_types['reminder_3']} Reminder-3 (4 days after)")
        if 'last_reminder' in reminder_types:
            summary_parts.append(f"{reminder_types['last_reminder']} Last Reminder (7 days after)")
        if 'discontinuation' in reminder_types:
            summary_parts.append(f"{reminder_types['discontinuation']} Discontinuation (10 days after)")
        
        summary = f"Found {len(reminder_installments)} installments requiring reminders:\n" + "\n".join(summary_parts)
        
        return JsonResponse({
            'success': True,
            'message': summary,
            'reminders_count': len(reminder_installments),
            'reminder_types': reminder_types,
            'script_info': 'Use the installment reminder script or batch file to send these reminders.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error checking installment reminders: {str(e)}'
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def mark_installment_paid(request):
    """Mark an installment as paid"""
    try:
        data = json.loads(request.body)
        installment_id = data.get('installment_id')
        registration_no = data.get('registration_no')
        
        from core.models import FeeInstallments
        from datetime import date
        
        # Get the installment
        installment = FeeInstallments.objects.get(installment_id=installment_id)
        
        # Mark as paid
        installment.status = 'Paid'
        installment.paid_date = date.today()
        installment.save()
        
        # Update fee details
        try:
            fee_details = FeeDetails.objects.get(registration_no=installment.registration_no)
            fee_details.fees_remaining = max(0, fee_details.fees_remaining - installment.amount)
            fee_details.save()
        except FeeDetails.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Installment {installment.installment_no} marked as paid successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error marking installment as paid: {str(e)}'
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_installment_records(request):
    """Create installment records for legacy students"""
    try:
        data = json.loads(request.body)
        registration_no = data.get('registration_no')
        
        student = Student.objects.get(registration_no=registration_no)
        fee_details = FeeDetails.objects.get(registration_no=student)
        
        from core.models import FeeInstallments
        from datetime import date, timedelta
        
        # Check if installments already exist
        existing_installments = FeeInstallments.objects.filter(registration_no=student).count()
        if existing_installments > 0:
            return JsonResponse({
                'success': False,
                'message': 'Installment records already exist for this student'
            })
        
        # Create installment records based on fee details
        installments_created = 0
        today = date.today()
        
        # Create installments based on number_of_installments
        for i in range(1, fee_details.number_of_installments + 1):
            amount = fee_details.fees_per_installment
            due_date = today + timedelta(days=30 * i)  # Monthly intervals
            
            # Check if specific installment data exists in fee_details
            if i == 1 and fee_details.first_installment:
                amount = fee_details.first_installment
                if fee_details.first_installment_date:
                    due_date = fee_details.first_installment_date
            elif i == 2 and fee_details.second_installment:
                amount = fee_details.second_installment
                if fee_details.second_installment_date:
                    due_date = fee_details.second_installment_date
            elif i == 3 and fee_details.third_installment:
                amount = fee_details.third_installment
                if fee_details.third_installment_date:
                    due_date = fee_details.third_installment_date
            
            FeeInstallments.objects.create(
                registration_no=student,
                installment_no=i,
                amount=amount,
                due_date=due_date,
                status='Due'
            )
            installments_created += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {installments_created} installment records!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating installment records: {str(e)}'
        })


@login_required
def send_student_whatsapp(request, registration_no):
    """Send WhatsApp reminder to a specific student"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Only POST method allowed'})
    
    try:
        import subprocess
        import os
        
        # Get student details
        student = Student.objects.get(registration_no=registration_no)
        
        # Create a temporary script to send WhatsApp to specific student
        script_content = f'''
import os
import sys
import django
from datetime import date, datetime

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fees_management_system.settings')
django.setup()

from core.models import Student, FeeInstallments, ParentInfo
import pywhatkit as pwk
import time

# Configuration
ADMIN_NUMBER = '+919822574252'

def send_whatsapp_message(phone_number, message):
    try:
        if not phone_number.startswith('+'):
            phone_number = '+91' + phone_number.replace('+91', '').replace(' ', '').replace('-', '')
        
        pwk.sendwhatmsg_instantly(phone_number, message, wait_time=10, tab_close=True)
        time.sleep(2)
        return True
    except Exception as e:
        print(f"Error sending WhatsApp message to {{phone_number}}: {{e}}")
        return False

def send_admin_notification(message_type, student_name, phone_number, message_sent):
    try:
        admin_message = f"""📱 WhatsApp Message Sent - Pillay Sir's ICSE Classes

Message Type: {{message_type.upper()}}
Student: {{student_name}}
Sent To: {{phone_number}}
Timestamp: {{datetime.now().strftime('%d %B %Y, %I:%M %p')}}

Message Content:
{{message_sent}}

---
This is an automated notification from the WhatsApp Reminder System."""
        
        return send_whatsapp_message(ADMIN_NUMBER, admin_message)
    except Exception as e:
        print(f"Error sending admin notification: {{e}}")
        return False

# Get specific student
try:
    student = Student.objects.get(registration_no={registration_no})
    print(f"Processing student: {{student.full_name}}")
    
    # Check for pending installments
    pending_installments = FeeInstallments.objects.filter(
        registration_no=student,
        status__in=['Due', 'Pending']
    ).order_by('due_date')
    
    if not pending_installments.exists():
        print("No pending installments found for this student.")
        exit()
    
    # Get the most urgent installment
    installment = pending_installments.first()
    today = date.today()
    days_difference = (today - installment.due_date).days
    
    # Determine message type
    if days_difference <= -1:
        message = f"Pillay Sir's ICSE Classes - Payment Reminder\\n\\nDear Parent,\\n\\nThis is to remind you that your ward {{student.full_name}}'s installment of ₹{{installment.amount}} is due on {{installment.due_date.strftime('%d %B %Y')}}.\\n\\nPlease make the payment at your earliest convenience.\\n\\nThank you,\\nPillay Sir"
        message_type = "upcoming_payment"
    elif days_difference >= 0:
        message = f"Pillay Sir's ICSE Classes - Overdue Payment\\n\\nDear Parent,\\n\\nYour ward {{student.full_name}}'s installment of ₹{{installment.amount}} was due on {{installment.due_date.strftime('%d %B %Y')}} and is now overdue.\\n\\nPlease make the payment immediately to avoid any inconvenience.\\n\\nThank you,\\nPillay Sir"
        message_type = "overdue_payment"
    
    # Get parent contact info
    try:
        parent_info = student.parentinfo
        phone_numbers = []
        
        if parent_info.father_mobile:
            phone_numbers.append(('Father', parent_info.father_mobile))
        if parent_info.mother_mobile:
            phone_numbers.append(('Mother', parent_info.mother_mobile))
        
        if not phone_numbers:
            print("No phone numbers found for this student.")
            exit()
        
        # Send messages
        for parent_type, phone_number in phone_numbers:
            print(f"Sending message to {{parent_type}}: {{phone_number}}")
            success = send_whatsapp_message(phone_number, message)
            
            if success:
                print(f"✅ Message sent successfully to {{parent_type}}")
                # Send admin notification
                send_admin_notification(message_type, student.full_name, phone_number, message)
            else:
                print(f"❌ Failed to send message to {{parent_type}}")
        
        print("WhatsApp sending process completed!")
        
    except Exception as e:
        print(f"Error getting parent info: {{e}}")
        
except Student.DoesNotExist:
    print(f"Student with registration number {registration_no} not found.")
except Exception as e:
    print(f"Error: {{e}}")
'''
        
        # Write temporary script
        temp_script_path = os.path.join(os.path.dirname(__file__), '..', 'temp_whatsapp_single.py')
        with open(temp_script_path, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        # Run the script
        result = subprocess.run([
            'python', temp_script_path
        ], capture_output=True, text=True, cwd=os.path.dirname(temp_script_path))
        
        # Clean up temporary script
        try:
            os.remove(temp_script_path)
        except:
            pass
        
        if result.returncode == 0:
            return JsonResponse({
                'success': True,
                'message': f'WhatsApp reminder sent successfully to {student.full_name}!',
                'output': result.stdout
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'Error sending WhatsApp: {result.stderr}',
                'output': result.stdout
            })
            
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'Student with registration number {registration_no} not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error sending WhatsApp: {str(e)}'
        })


@login_required
def generate_receipt(request, registration_no):
    """Generate receipt for a student"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.http import HttpResponse
        from datetime import datetime
        import io
        
        # Get student and fee details
        student = Student.objects.get(registration_no=registration_no)
        fee_details = FeeDetails.objects.get(registration_no=student)
        parent_info = ParentInfo.objects.get(registration_no=student)
        academic_info = AcademicInfo.objects.get(registration_no=student)
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            alignment=TA_LEFT,
            textColor=colors.darkblue
        )
        
        # Title
        title = Paragraph("🎓 Pillay Sir's ICSE Classes", title_style)
        subtitle = Paragraph("Fee Receipt", styles['Heading2'])
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 20))
        
        # Receipt details
        receipt_data = [
            ['Receipt Date:', datetime.now().strftime('%d %B %Y')],
            ['Receipt No:', f'RCP-{registration_no}-{datetime.now().strftime("%Y%m%d")}'],
        ]
        
        receipt_table = Table(receipt_data, colWidths=[2*inch, 3*inch])
        receipt_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(receipt_table)
        elements.append(Spacer(1, 20))
        
        # Student Information
        elements.append(Paragraph("Student Information", header_style))
        student_data = [
            ['Registration No:', str(student.registration_no)],
            ['Student Name:', student.full_name],
            ['Course:', academic_info.enrolled_course],
            ['Branch:', str(academic_info.branch_code)],
            ['Father\'s Name:', parent_info.father_name],
            ['Mother\'s Name:', parent_info.mother_name],
            ['Contact:', parent_info.father_mobile or parent_info.mother_mobile or 'N/A'],
        ]
        
        student_table = Table(student_data, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 20))
        
        # Fee Details
        elements.append(Paragraph("Fee Details", header_style))
        fees_paid = fee_details.total_fees - fee_details.fees_remaining
        
        fee_data = [
            ['Description', 'Amount (₹)'],
            ['Total Fees', f'{fee_details.total_fees:,.2f}'],
            ['Fees Paid', f'{fees_paid:,.2f}'],
            ['Fees Remaining', f'{fee_details.fees_remaining:,.2f}'],
            ['Number of Installments', str(fee_details.number_of_installments)],
        ]
        
        fee_table = Table(fee_data, colWidths=[3*inch, 2*inch])
        fee_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(fee_table)
        elements.append(Spacer(1, 30))
        
        # Installment Details (if available)
        try:
            installments = FeeInstallments.objects.filter(registration_no=student).order_by('installment_no')
            if installments.exists():
                elements.append(Paragraph("Installment Details", header_style))
                
                installment_data = [['Installment', 'Amount (₹)', 'Due Date', 'Status']]
                for inst in installments:
                    installment_data.append([
                        f'Installment {inst.installment_no}',
                        f'{inst.amount:,.2f}',
                        inst.due_date.strftime('%d %b %Y'),
                        inst.status
                    ])
                
                installment_table = Table(installment_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                installment_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(installment_table)
                elements.append(Spacer(1, 20))
        except:
            pass
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_text = """
        <para align="center">
        <b>Thank you for choosing Pillay Sir's ICSE Classes!</b><br/>
        For any queries, please contact us.<br/>
        <i>This is a computer-generated receipt.</i>
        </para>
        """
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Receipt_{student.full_name}_{registration_no}.pdf"'
        
        return response
        
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': f'Student with registration number {registration_no} not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error generating receipt: {str(e)}'
        })


@login_required
def whatsapp_message_center(request):
    """WhatsApp Message Center - categorized students for different reminder types"""
    from datetime import date
    
    today = date.today()
    
    # Get birthday students
    birthday_students = Student.objects.filter(
        date_of_birth__month=today.month,
        date_of_birth__day=today.day
    ).select_related('parentinfo', 'academicinfo', 'feedetails')
    
    # Initialize reminder categories
    reminder_categories = {
        'reminder_1': [],  # -3 days before due
        'reminder_2': [],  # +1 day after due
        'reminder_3': [],  # +4 days after due
        'last_reminder': [],  # +7 days after due
        'final_notice': []  # +10 days after due
    }
    
    # Get all non-paid installments
    all_installments = FeeInstallments.objects.exclude(
        status='Paid'
    ).select_related('registration_no', 'registration_no__parentinfo', 'registration_no__academicinfo')
    
    # Categorize installments by reminder type
    for installment in all_installments:
        days_difference = (today - installment.due_date).days
        
        if days_difference == -3:
            reminder_categories['reminder_1'].append(installment)
        elif days_difference == 1:
            reminder_categories['reminder_2'].append(installment)
        elif days_difference == 4:
            reminder_categories['reminder_3'].append(installment)
        elif days_difference == 7:
            reminder_categories['last_reminder'].append(installment)
        elif days_difference == 10:
            reminder_categories['final_notice'].append(installment)
    
    # Calculate totals
    total_reminders = sum(len(category) for category in reminder_categories.values())
    total_birthday_messages = birthday_students.count()
    
    context = {
        'birthday_students': birthday_students,
        'reminder_categories': reminder_categories,
        'total_reminders': total_reminders,
        'total_birthday_messages': total_birthday_messages,
        'total_messages': total_reminders + total_birthday_messages,
        'today': today,
    }
    
    return render(request, 'core/whatsapp_message_center.html', context)


@login_required
@csrf_exempt
def send_individual_whatsapp(request):
    """Send individual WhatsApp message"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        data = json.loads(request.body)
        message_type = data.get('message_type')
        student_id = data.get('student_id')
        installment_id = data.get('installment_id')
        
        # Import the message sending functionality
        from send_whatsapp_reminders_installments import send_whatsapp_message, MESSAGE_TEMPLATES, format_message, send_admin_notification
        
        if message_type == 'birthday':
            # Handle birthday message
            student = Student.objects.get(registration_no=student_id)
            message = format_message(MESSAGE_TEMPLATES['birthday'], student=student)
            
            # Get parent contact info
            parent_info = student.parentinfo
            phone_numbers = []
            
            if parent_info.father_mobile:
                phone_numbers.append(('Father', parent_info.father_mobile))
            if parent_info.mother_mobile:
                phone_numbers.append(('Mother', parent_info.mother_mobile))
            
            if not phone_numbers:
                return JsonResponse({'success': False, 'message': 'No phone numbers found for this student'})
            
            # Send to each parent
            sent_count = 0
            for parent_type, phone_number in phone_numbers:
                success = send_whatsapp_message(phone_number, message)
                
                # Create analytics record
                analytics = WhatsAppMessageAnalytics.objects.create(
                    student=student,
                    message_type='birthday',
                    message_content=message,
                    phone_number=phone_number,
                    parent_type=parent_type,
                    delivery_status='SENT' if success else 'FAILED'
                )
                
                if success:
                    sent_count += 1
                    # Send admin notification
                    send_admin_notification('birthday', student.full_name, phone_number, message)
            
            return JsonResponse({
                'success': True,
                'message': f'Birthday message sent to {sent_count} contact(s)',
                'sent_count': sent_count
            })
        
        else:
            # Handle fee reminder message
            installment = FeeInstallments.objects.get(installment_id=installment_id)
            
            # Get message template
            template = MESSAGE_TEMPLATES.get(message_type)
            if not template:
                return JsonResponse({'success': False, 'message': 'Invalid message type'})
            
            message = format_message(template, installment)
            
            # Get parent contact info
            parent_info = installment.registration_no.parentinfo
            phone_numbers = []
            
            if parent_info.father_mobile:
                phone_numbers.append(('Father', parent_info.father_mobile))
            if parent_info.mother_mobile:
                phone_numbers.append(('Mother', parent_info.mother_mobile))
            
            if not phone_numbers:
                return JsonResponse({'success': False, 'message': 'No phone numbers found for this student'})
            
            # Send to each parent
            sent_count = 0
            for parent_type, phone_number in phone_numbers:
                success = send_whatsapp_message(phone_number, message)
                
                # Create analytics record
                analytics = WhatsAppMessageAnalytics.objects.create(
                    student=installment.registration_no,
                    installment=installment,
                    message_type=message_type,
                    message_content=message,
                    phone_number=phone_number,
                    parent_type=parent_type,
                    delivery_status='SENT' if success else 'FAILED'
                )
                
                if success:
                    sent_count += 1
                    # Send admin notification
                    send_admin_notification(message_type, installment.registration_no.full_name, phone_number, message)
            
            return JsonResponse({
                'success': True,
                'message': f'Reminder sent to {sent_count} contact(s)',
                'sent_count': sent_count
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error sending message: {str(e)}'})


@login_required
@csrf_exempt
def send_bulk_whatsapp_messages(request):
    """Send all WhatsApp messages (birthdays + reminders) in bulk"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        from datetime import date
        from send_whatsapp_reminders_installments import (
            send_whatsapp_message, MESSAGE_TEMPLATES, format_message, 
            send_admin_notification, get_birthday_students, get_installments_for_reminders
        )
        
        today = date.today()
        sent_messages = []
        failed_messages = []
        
        # Get birthday students
        birthday_students = get_birthday_students()
        
        # Send birthday messages
        for student in birthday_students:
            try:
                birthday_message = format_message(MESSAGE_TEMPLATES['birthday'], student=student)
                
                # Get parent contact info
                parent_info = student.parentinfo
                phone_numbers = []
                
                if parent_info.father_mobile:
                    phone_numbers.append(('Father', parent_info.father_mobile))
                if parent_info.mother_mobile:
                    phone_numbers.append(('Mother', parent_info.mother_mobile))
                
                # Send to each parent
                for parent_type, phone_number in phone_numbers:
                    success = send_whatsapp_message(phone_number, birthday_message)
                    if success:
                        sent_messages.append({
                            'type': 'birthday',
                            'student': student.full_name,
                            'phone': phone_number,
                            'parent_type': parent_type
                        })
                        send_admin_notification('birthday', student.full_name, phone_number, birthday_message)
                    else:
                        failed_messages.append({
                            'type': 'birthday',
                            'student': student.full_name,
                            'phone': phone_number,
                            'error': 'Send failed'
                        })
            except Exception as e:
                failed_messages.append({
                    'type': 'birthday',
                    'student': student.full_name,
                    'error': str(e)
                })
        
        # Get installments for reminders
        installments = get_installments_for_reminders()
        
        # Send fee reminders
        for installment in installments:
            try:
                # Determine message type based on days difference
                days_difference = (today - installment.due_date).days
                message_type = None
                
                if days_difference == -3:
                    message_type = 'reminder_1'
                elif days_difference == 1:
                    message_type = 'reminder_2'
                elif days_difference == 4:
                    message_type = 'reminder_3'
                elif days_difference == 7:
                    message_type = 'last_reminder'
                elif days_difference == 10:
                    message_type = 'discontinuation'
                
                if message_type and message_type in MESSAGE_TEMPLATES:
                    message = format_message(MESSAGE_TEMPLATES[message_type], installment)
                    
                    # Get parent contact info
                    parent_info = installment.registration_no.parentinfo
                    phone_numbers = []
                    
                    if parent_info.father_mobile:
                        phone_numbers.append(('Father', parent_info.father_mobile))
                    if parent_info.mother_mobile:
                        phone_numbers.append(('Mother', parent_info.mother_mobile))
                    
                    # Send to each parent
                    for parent_type, phone_number in phone_numbers:
                        success = send_whatsapp_message(phone_number, message)
                        if success:
                            sent_messages.append({
                                'type': message_type,
                                'student': installment.registration_no.full_name,
                                'phone': phone_number,
                                'parent_type': parent_type,
                                'installment': installment.installment_no
                            })
                            send_admin_notification(message_type, installment.registration_no.full_name, phone_number, message)
                        else:
                            failed_messages.append({
                                'type': message_type,
                                'student': installment.registration_no.full_name,
                                'phone': phone_number,
                                'error': 'Send failed'
                            })
            except Exception as e:
                failed_messages.append({
                    'type': 'reminder',
                    'student': installment.registration_no.full_name,
                    'error': str(e)
                })
        
        # Send admin summary
        try:
            summary_message = f"""Pillay Sir's ICSE Classes - WhatsApp Summary

🎂 Birthdays: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'birthday']) or '—'}
Reminder 1: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'reminder_1']) or '—'}
Reminder 2: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'reminder_2']) or '—'}
Reminder 3: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'reminder_3']) or '—'}
Last Reminder: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'last_reminder']) or '—'}
Final Notice: {', '.join([msg['student'] for msg in sent_messages if msg['type'] == 'discontinuation']) or '—'}

Total Sent: {len(sent_messages)}
Failed: {len(failed_messages)}"""
            
            admin_number = '+919822574252'  # Replace with actual admin number
            send_whatsapp_message(admin_number, summary_message)
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Bulk messages sent successfully. {len(sent_messages)} sent, {len(failed_messages)} failed.',
            'sent_count': len(sent_messages),
            'failed_count': len(failed_messages),
            'sent_messages': sent_messages,
            'failed_messages': failed_messages
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error sending bulk messages: {str(e)}'})


@login_required
def message_analytics_dashboard(request):
    """Message Analytics Dashboard with delivery tracking, response analytics, and success rates"""
    from datetime import date, timedelta
    from django.db.models import Avg, Count, Q
    
    # Date range for analytics (last 30 days by default)
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    
    # Get date range from request if provided
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    # Get analytics data for the date range
    analytics = WhatsAppMessageAnalytics.objects.filter(
        sent_timestamp__date__range=[start_date, end_date]
    ).select_related('student', 'installment')
    
    # Overall Statistics
    total_messages = analytics.count()
    delivered_messages = analytics.filter(delivery_status='DELIVERED').count()
    read_messages = analytics.filter(delivery_status='READ').count()
    failed_messages = analytics.filter(delivery_status='FAILED').count()
    responded_messages = analytics.exclude(response_type='NONE').count()
    
    # Calculate rates
    delivery_rate = (delivered_messages / total_messages * 100) if total_messages > 0 else 0
    read_rate = (read_messages / total_messages * 100) if total_messages > 0 else 0
    response_rate = (responded_messages / total_messages * 100) if total_messages > 0 else 0
    
    # Message Type Performance
    message_type_stats = []
    for msg_type, msg_label in WhatsAppMessageAnalytics.MESSAGE_TYPE_CHOICES:
        type_analytics = analytics.filter(message_type=msg_type)
        type_count = type_analytics.count()
        if type_count > 0:
            type_delivered = type_analytics.filter(delivery_status='DELIVERED').count()
            type_read = type_analytics.filter(delivery_status='READ').count()
            type_responded = type_analytics.exclude(response_type='NONE').count()
            
            message_type_stats.append({
                'type': msg_type,
                'label': msg_label,
                'count': type_count,
                'delivery_rate': (type_delivered / type_count * 100) if type_count > 0 else 0,
                'read_rate': (type_read / type_count * 100) if type_count > 0 else 0,
                'response_rate': (type_responded / type_count * 100) if type_count > 0 else 0,
            })
    
    # Daily Performance (last 7 days)
    daily_stats = []
    for i in range(7):
        target_date = end_date - timedelta(days=i)
        day_analytics = analytics.filter(sent_timestamp__date=target_date)
        day_count = day_analytics.count()
        day_delivered = day_analytics.filter(delivery_status='DELIVERED').count()
        day_responded = day_analytics.exclude(response_type='NONE').count()
        
        daily_stats.append({
            'date': target_date,
            'messages_sent': day_count,
            'delivery_rate': (day_delivered / day_count * 100) if day_count > 0 else 0,
            'response_rate': (day_responded / day_count * 100) if day_count > 0 else 0,
        })
    
    daily_stats.reverse()  # Show oldest to newest
    
    # Response Type Distribution
    response_stats = []
    for resp_type, resp_label in WhatsAppMessageAnalytics.RESPONSE_TYPE_CHOICES:
        if resp_type != 'NONE':
            resp_count = analytics.filter(response_type=resp_type).count()
            if resp_count > 0:
                response_stats.append({
                    'type': resp_type,
                    'label': resp_label,
                    'count': resp_count,
                    'percentage': (resp_count / responded_messages * 100) if responded_messages > 0 else 0
                })
    
    # Top Engaged Students
    engaged_students = analytics.filter(is_engaged=True).values(
        'student__first_name', 'student__last_name', 'student__registration_no'
    ).annotate(
        engagement_score=Avg('engagement_score'),
        message_count=Count('analytics_id'),
        response_count=Count('analytics_id', filter=~Q(response_type='NONE'))
    ).order_by('-engagement_score')[:10]
    
    # Recent Message History (last 50 messages)
    recent_messages = analytics.order_by('-sent_timestamp')[:50]
    
    # Average Response Times by Message Type
    response_time_stats = []
    for msg_type, msg_label in WhatsAppMessageAnalytics.MESSAGE_TYPE_CHOICES:
        type_responses = analytics.filter(
            message_type=msg_type,
            response_time_seconds__isnull=False
        )
        if type_responses.exists():
            avg_response_time = type_responses.aggregate(
                avg_time=Avg('response_time_seconds')
            )['avg_time']
            
            response_time_stats.append({
                'type': msg_type,
                'label': msg_label,
                'avg_response_hours': avg_response_time / 3600 if avg_response_time else 0,
                'response_count': type_responses.count()
            })
    
    # Generate daily summaries for the date range
    summaries = []
    current_date = start_date
    while current_date <= end_date:
        summary = MessageAnalyticsSummary.generate_daily_summary(current_date)
        if summary:
            summaries.append(summary)
        current_date += timedelta(days=1)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_messages': total_messages,
        'delivered_messages': delivered_messages,
        'read_messages': read_messages,
        'failed_messages': failed_messages,
        'responded_messages': responded_messages,
        'delivery_rate': round(delivery_rate, 1),
        'read_rate': round(read_rate, 1),
        'response_rate': round(response_rate, 1),
        'message_type_stats': message_type_stats,
        'daily_stats': daily_stats,
        'response_stats': response_stats,
        'engaged_students': engaged_students,
        'recent_messages': recent_messages,
        'response_time_stats': response_time_stats,
        'summaries': summaries,
    }
    
    return render(request, 'core/message_analytics_dashboard.html', context)


@login_required
@csrf_exempt
def update_message_status(request):
    """Update message delivery status and response information"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        data = json.loads(request.body)
        analytics_id = data.get('analytics_id')
        delivery_status = data.get('delivery_status')
        response_type = data.get('response_type')
        response_content = data.get('response_content', '')
        
        analytics = WhatsAppMessageAnalytics.objects.get(analytics_id=analytics_id)
        
        # Update delivery status
        if delivery_status:
            analytics.delivery_status = delivery_status
            
            # Set timestamps based on status
            now = datetime.now()
            if delivery_status == 'DELIVERED' and not analytics.delivery_timestamp:
                analytics.delivery_timestamp = now
                if analytics.sent_timestamp:
                    analytics.delivery_time_seconds = int((now - analytics.sent_timestamp).total_seconds())
            
            elif delivery_status == 'READ' and not analytics.read_timestamp:
                analytics.read_timestamp = now
                if analytics.sent_timestamp:
                    analytics.read_time_seconds = int((now - analytics.sent_timestamp).total_seconds())
        
        # Update response information
        if response_type:
            analytics.response_type = response_type
            analytics.response_content = response_content
            
            if not analytics.response_timestamp:
                analytics.response_timestamp = datetime.now()
                if analytics.sent_timestamp:
                    analytics.response_time_seconds = int((analytics.response_timestamp - analytics.sent_timestamp).total_seconds())
        
        # Recalculate engagement score
        analytics.calculate_engagement_score()
        analytics.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Message status updated successfully',
            'engagement_score': analytics.engagement_score
        })
    
    except WhatsAppMessageAnalytics.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Message analytics record not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error updating status: {str(e)}'})


@login_required
def export_analytics_data(request):
    """Export analytics data to Excel"""
    try:
        import pandas as pd
        from datetime import date, timedelta
        
        # Get date range
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        
        if request.GET.get('start_date'):
            start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        if request.GET.get('end_date'):
            end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
        
        # Get analytics data
        analytics = WhatsAppMessageAnalytics.objects.filter(
            sent_timestamp__date__range=[start_date, end_date]
        ).select_related('student', 'installment')
        
        # Prepare data for export
        export_data = []
        for record in analytics:
            export_data.append({
                'Date': record.sent_timestamp.strftime('%Y-%m-%d'),
                'Time': record.sent_timestamp.strftime('%H:%M:%S'),
                'Student Name': record.student.full_name,
                'Registration No': record.student.registration_no,
                'Message Type': record.get_message_type_display(),
                'Phone Number': record.phone_number,
                'Parent Type': record.parent_type,
                'Delivery Status': record.get_delivery_status_display(),
                'Response Type': record.get_response_type_display(),
                'Delivery Time': record.delivery_time_formatted,
                'Response Time': record.response_time_formatted,
                'Engagement Score': record.engagement_score,
                'Is Engaged': 'Yes' if record.is_engaged else 'No',
            })
        
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="message_analytics_{start_date}_to_{end_date}.xlsx"'
        
        # Write to Excel
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Message Analytics', index=False)
            
            # Add summary sheet
            summary_data = {
                'Metric': ['Total Messages', 'Delivered', 'Read', 'Failed', 'Responses', 'Delivery Rate (%)', 'Read Rate (%)', 'Response Rate (%)'],
                'Value': [
                    len(export_data),
                    len([d for d in export_data if d['Delivery Status'] == 'Delivered']),
                    len([d for d in export_data if d['Delivery Status'] == 'Read']),
                    len([d for d in export_data if d['Delivery Status'] == 'Failed']),
                    len([d for d in export_data if d['Response Type'] != 'No Response']),
                    round(len([d for d in export_data if d['Delivery Status'] == 'Delivered']) / len(export_data) * 100, 1) if export_data else 0,
                    round(len([d for d in export_data if d['Delivery Status'] == 'Read']) / len(export_data) * 100, 1) if export_data else 0,
                    round(len([d for d in export_data if d['Response Type'] != 'No Response']) / len(export_data) * 100, 1) if export_data else 0,
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        return response
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error exporting data: {str(e)}'})
